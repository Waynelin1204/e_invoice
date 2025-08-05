# ====== Python 標準函式庫 ======
import os
import json
import logging
import subprocess
from io import BytesIO
from datetime import datetime, timedelta
from collections import defaultdict
import xml.etree.ElementTree as ET
from django.utils.timezone import localtime
from decimal import Decimal, InvalidOperation


# ====== 第三方套件 ======
import pandas as pd
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# ====== Django 基礎功能 ======
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.utils import timezone
from django.urls import reverse

# ====== Django DB 操作 ======
from django.db import connection
from django.db import transaction
from django.db.models import Q, Count, Sum

# ====== Django 使用者與驗證 ======
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test

# ====== 專案內部（Model 與 Form） ======
from e_invoices.models import (
    RegisterForm, LoginForm,
    Twa0101, Twa0101Item, Ocr, Ocritem, Company, UserProfile,TWAllowance,TWAllowanceLineItem,
    NumberDistribution, TWB2BMainItem, TWB2BLineItem, CompanyNotificationConfig, SystemConfig
)
from e_invoices.forms import NumberDistributionForm
from e_invoices.services.validate_allowance import validate_allowance
from decimal import Decimal, InvalidOperation
from e_invoices.services import generate_G0401_xml_files, generate_G0501_xml_files, generate_allowance_pdf, send_allowance_summary_email, send_allowance_canceled_email, send_allowance_deleted_email


def update_decimal_field(obj, field_name, raw_val):
    """
    如果 raw_val 有值且為合法 Decimal，更新 obj 的指定欄位。
    否則保留原值不動。
    """
    val = (raw_val or '').strip()
    if val:
        try:
            setattr(obj, field_name, Decimal(val))
        except InvalidOperation:
            # 你可以改為 log 錯誤或顯示錯誤訊息
            pass

@login_required
def twallowance(request):
    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile
    
    # 取得該使用者可查看的公司名稱列表
    viewable_company_codes = user_profile.viewable_companies.values_list('company_id', flat=True)
    # 取得B2B或B2C列表
    b2b_b2c_filter =  request.GET.get("b2b_b2c")
    
    # 計算從今天開始往前推的60天的日期
    sixty_days_ago = datetime.today() - timedelta(days=60)
    
    # 篩選條件：只顯示最近60天的發票
    filter_conditions = Q(company__in=viewable_company_codes) & Q(erp_date__gte=sixty_days_ago) & Q(b2b_b2c=b2b_b2c_filter)

    # 查詢符合條件的資料，並使用 prefetch_related 來查詢發票明細
    allowances = TWAllowance.objects.filter(filter_conditions).order_by('-erp_date').prefetch_related('items__linked_invoice')
    company_options = user_profile.viewable_companies.all()
    
    # 分頁：每頁顯示25筆資料
    paginator = Paginator(allowances, 25)  # 每頁顯示25筆資料
    page_number = request.GET.get('page')  # 取得當前頁數
    page_obj = paginator.get_page(page_number)  # 根據頁數取得對應的資料
    validation_results = validate_allowance(page_obj)

    # 驗證每個折讓單
    #validated_allowances = []
    for result in validation_results:
        allowance = result["allowance"]
        allowance.is_allowance_valid = result["is_allowance_valid"]
        allowance.is_valid_amount = result["is_valid_amount"]
        allowance.is_valid_tax = result["is_valid_tax"]
        # 若需要也可以掛載 validated_items
        allowance.validated_items = result["validated_items"]
        #allowance.invalid_invoice_numbers = result["invalid_invoice_numbers"]
    
        
    
        print(f"\n📄 折讓單號：{getattr(allowance, 'allowance_number', allowance.id)}")
        print(f"   ✅ 是否有效折讓單：{allowance.is_allowance_valid}")
        print(f"   📌 金額驗證：{allowance.is_valid_amount}")
        print(f"   📌 稅額驗證：{allowance.is_valid_tax}")
        
        print("   ── 明細驗證 ──")
        for item_result in result["validated_items"]:
            item = item_result["item"]
            print(f"      - 明細ID: {item.id}, is_valid_amount: {item_result['is_valid_amount']}, is_valid_tax: {item_result['is_valid_tax']}")
        
        print("   ── 發票剩餘可折讓金額與稅額 ──")
        for invoice_number, amount in result.get("invoice_remaining_amounts", {}).items():
            tax = result.get("invoice_remaining_taxes", {}).get(invoice_number, "N/A")
            print(f"      - 發票號碼：{invoice_number}, 剩餘金額：{amount}, 剩餘稅額：{tax}")



    # 傳遞資料給模板
    context = {
        "page_obj": page_obj,
        "company_options": company_options,
        "allowances":  [r['allowance'] for r in validation_results],
    }


    return render(request, 'twallowance.html', context)

#======================================================B2B發票篩選=======================================================

def twallowance_filter(request):
    # 預設顯示筆數
    display_limit = int(request.GET.get("display_limit", 20))  # 默認為20筆資料

    # 其他篩選條件
    allowance_status_filter = request.GET.get("allowance_status")
    line_tax_type_filter = request.GET.get("line_tax_type")
    company_id_filter = request.GET.get("company_id")  # 新增公司篩選條件
    b2b_b2c_filter =  request.GET.get("b2b_b2c")


    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile

    # 取得該使用者可查看的公司名稱列表
    company_options = user_profile.viewable_companies.all()
    viewable_company_codes = user_profile.viewable_companies.values_list('company_id', flat=True)


    # 計算兩個月前的日期
    two_months_ago = datetime.today() - timedelta(days=60)

    # 取得時間範圍的過濾條件（默認從兩個月前開始）
    start_date = request.GET.get("start_date", two_months_ago.strftime('%Y-%m-%d'))
    end_date = request.GET.get("end_date", datetime.today().strftime('%Y-%m-%d'))

    # 轉換成 datetime 格式
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    except ValueError:
        start_date = two_months_ago  # 如果格式錯誤，使用預設的兩個月前日期

    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        end_date = datetime.today()  # 如果格式錯誤，使用今天的日期

    if (end_date - start_date).days > 60:
        messages.error(request, "查詢區間不得超過 60 天")
        return redirect('twallowance')
    if start_date > end_date:
        messages.error(request, "開始日期不能大於結束日期")
        return redirect('twallowance')
    
    # 查詢條件構建
    filters = Q()
    if allowance_status_filter:
        filters &= Q(allowance_status=allowance_status_filter)
    if line_tax_type_filter:
        filters &= Q(line_tax_type=line_tax_type_filter)
    if company_id_filter:
        filters &= Q(company__company_id=company_id_filter)
    if b2b_b2c_filter:
        filters &= Q(b2b_b2c=b2b_b2c_filter)

    # 限制在兩個月前到今天的時間範圍內
    filters &= Q(erp_date__range=[start_date, end_date])

    # 加入公司權限過濾條件
    filters &= Q(company__in=viewable_company_codes)

    # 查詢所有符合條件的發票資料
    allowances_list = TWAllowance.objects.filter(filters).order_by('-erp_date')

    # 獲取篩選條件的選項
    allowance_status = TWAllowance.objects.values_list('allowance_status', flat=True).distinct()
    b2b_b2c = TWAllowance.objects.values_list('b2b_b2c', flat=True).distinct()

    # 分頁
    paginator = Paginator(allowances_list, display_limit)  # 每頁顯示的資料筆數
    page_number = request.GET.get('page')  # 獲取當前頁碼
    page_obj = paginator.get_page(page_number)  # 根據頁碼獲取相應的頁面資料


    # 分頁：每頁顯示25筆資料

    validation_results = validate_allowance(page_obj)

    # 驗證每個折讓單
    #validated_allowances = []
    for result in validation_results:
        allowance = result["allowance"]
        allowance.is_allowance_valid = result["is_allowance_valid"]
        allowance.is_valid_amount = result["is_valid_amount"]
        allowance.is_valid_tax = result["is_valid_tax"]
        # 若需要也可以掛載 validated_items
        allowance.validated_items = result["validated_items"]
        #allowance.invalid_invoice_numbers = result["invalid_invoice_numbers"]
    

    # 檢查公司ID篩選是否有效
    if company_id_filter and company_id_filter not in viewable_company_codes:
        messages.error(request, "您無權限查看該公司資料")
        return redirect('twallowance')

    return render(request, "twallowance.html", {
        "company_id_filter": company_id_filter,
        "page_obj": page_obj,  # 傳遞分頁結果
        "company_options": company_options,
        "b2b_b2c": b2b_b2c,
        "invoice_status": allowance_status,
        "display_limit": display_limit,  # 傳遞選擇的筆數
        "start_date": start_date.strftime('%Y-%m-%d'),  # 顯示篩選的開始日期
        "end_date": end_date.strftime('%Y-%m-%d'),  # 顯示篩選的結束日期
        "allowances":  [r['allowance'] for r in validation_results],
    })

#====================================================== 折讓單明細 =======================================================
    
# def twallowance_detail(request, id):
#     allowance = get_object_or_404(TWAllowance, id=id)
#     items = allowance.items.all()  # 確保有正確查詢
#     return render(request, 'allowance/twallowance_detail.html', {'allowance': allowance, 'items': items})

def twallowance_detail(request, id):
    allowance = get_object_or_404(TWAllowance, id=id)

    # 只抓 linked_invoice 為 None 的項目進行 mapping，其它不動
    unmapped_items = allowance.items.filter(linked_invoice__isnull=True)
    invoice_numbers = [item.line_original_invoice_number for item in unmapped_items]

    # 對應 invoice_number => TWB2BMainItem instance
    invoice_map = {
        inv.invoice_number: inv
        for inv in TWB2BMainItem.objects.filter(invoice_number__in=invoice_numbers)
    }

    updated_items = []
    for item in unmapped_items:
        matched_invoice = invoice_map.get(item.line_original_invoice_number)
        if matched_invoice:
            item.linked_invoice = matched_invoice
            updated_items.append(item)

    # 批次更新，只更新未設定的資料
    if updated_items:
        TWAllowanceLineItem.objects.bulk_update(updated_items, ['linked_invoice'])

    # 查詢全部 items（含已 mapping 的）
    items = allowance.items.select_related('linked_invoice').all()

    return render(request, 'allowance/twallowance_detail.html', {
        'allowance': allowance,
        'items': items,
    })
@csrf_exempt
def twallowance_update(request, id):
    allowance = get_object_or_404(TWAllowance, id=id)
    if request.method == 'POST':
        # 處理主項目資料
        allowance_period = request.POST.get('invoice_period', '').strip()
        erp_reference = request.POST.get('erp_reference', '').strip()
        seller_bp_id = request.POST.get('seller_bp_id', '').strip()
        buyer_bp_id = request.POST.get('buyer_bp_id', '').strip()
        allowance_amount = request.POST.get('allowance_amount', allowance.allowance_amount)
        allowance_tax = request.POST.get('allowance_tax', allowance.allowance_tax)
        allowance_period = request.POST.get('allowance_period', allowance.allowance_period)
        allowance_cancel_reason = request.POST.get('allowance_cancel_reason', allowance.allowance_cancel_reason)
        allowance_cancel_remark = request.POST.get('allowance_cancel_remark', allowance.allowance_cancel_remark)
        # 更新主項目資料

        allowance.allowance_period = allowance_period
        allowance.erp_reference = erp_reference
        allowance.seller_bp_id = seller_bp_id
        allowance.buyer_bp_id = buyer_bp_id
        allowance.allowance_amount = allowance_amount
        allowance.allowance_tax = allowance_tax
        allowance.allowance_cancel_reason = allowance_cancel_reason
        allowance.allowance_cancel_remark = allowance_cancel_remark


        allowance.save()  # 保存主項目資料

        # 更新明細項目資料
        for item in allowance.items.all():
            update_decimal_field(item, 'line_quantity', request.POST.get(f'line_quantity_{item.id}'))
            update_decimal_field(item, 'line_unit_price', request.POST.get(f'line_unit_price_{item.id}'))
            update_decimal_field(item, 'line_allowance_amount', request.POST.get(f'line_allowance_amount_{item.id}'))
            update_decimal_field(item, 'line_allowance_tax', request.POST.get(f'line_allowance_tax_{item.id}'))
            #line_quantity = request.POST.get(f'line_quantity_{item.id}', '').strip()
            #line_unit_price =  Decimal(request.POST.get(f'line_unit_price_{item.id}', '').strip())
            # line_allowance_amount = request.POST.get(f'line_allowance_amount_{item.id}', '').strip()
            #line_allowance_tax = Decimal(request.POST.get(f'line_allowance_tax_{item.id}', '').strip())
            #line_allowance_amount = Decimal(request.POST.get(f'line_allowance_amount_{item.id}', '').strip())



            #item.line_quantity = line_quantity 
            #item.line_unit_price = line_unit_price
            #item.line_allowance_amount = line_allowance_amount
            #item.line_allowance_tax = line_allowance_tax

            item.save()  # 保存明細項目資料

        return redirect('twallowance_detail', id=allowance.id)  # 更新後重定向到發票詳情頁面

    return render(request, 'allowance/twallowance_detail.html', {'allowance': allowance})
#======================================================B2B發票匯出=======================================================

@csrf_exempt
def twallowance_export_invoices(request):
    config = SystemConfig.objects.first()
    to_email = config.operator_output_email_address if config else None
    F0401_XSD_path = config.F0401_XSD_path
    F0501_XSD_path = config.F0501_XSD_path
    G0401_XSD_path = config.G0401_XSD_path
    G0501_XSD_path = config.G0501_XSD_path
    A0201_XSD_path = config.A0201_XSD_path
    A0301_XSD_path = config.A0301_XSD_path
    A0101_XSD_path = config.A0101_XSD_path
    A0102_XSD_path = config.A0102_XSD_path
    A0202_XSD_path = config.A0202_XSD_path
    A0302_XSD_path = config.A0302_XSD_path
    B0101_XSD_path = config.B0101_XSD_path
    B0102_XSD_path = config.B0101_XSD_path
    B0201_XSD_path = config.B0201_XSD_path
    B0202_XSD_path = config.B0101_XSD_path
    if request.method != 'POST':
        #return HttpResponse("Only POST allowed", status=405)
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    return_querystring = request.POST.get('return_querystring', '')  # ⬅️ 關鍵步驟

    #if not selected_ids:
    #    return HttpResponse("No invoice IDs provided", status=400)

    if not selected_ids:
        messages.error(request, "開立失敗：未選擇任何折讓單。")
        redirect_url = reverse('twallowance')
        if return_querystring:
            redirect_url += '?' + return_querystring
        return redirect(redirect_url)
    


    #invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items')
    allowances = TWAllowance.objects.filter(id__in=selected_ids).exclude(allowance_status='已開立')
    #if not allowances.exists():
    #    return HttpResponse("No invoices found", status=404)
    if not allowances.exists():
        messages.error(request, "找不到折讓單，請確認是否重複或資料不完整")
        redirect_url = reverse('twallowance')
        if return_querystring:
            redirect_url += '?' + return_querystring
        return redirect(redirect_url)
    #all_selected_allowances = TWAllowance.objects.filter(id__in=selected_ids)

    #allowances = all_selected_allowances.exclude(allowance_status='已開立')

    #if not allowances.exists():
    #    return HttpResponse("No valid allowances found", status=404)    

    # # 先找出所有選取的發票
    # all_selected_invoices = TWAllowance.objects.filter(id__in=selected_ids)
    # # 計算「已開立」的數量
    # excluded_count = all_selected_invoices.filter(invoice_status='已開立').count()

    # # 篩選出尚未開立的發票
    # invoices = all_selected_invoices.exclude(invoice_status='已開立').prefetch_related('items')

    # # 如果有被排除的，就提示使用者
    # if excluded_count > 0:
    #     messages.warning(request, f"{excluded_count} 筆已開立的發票已排除，未匯出。")

    # # 1️⃣ 統計各公司所需發票數
    # invoice_count_by_company_code = defaultdict(int)
    # for invoice in invoices:
    #     #invoice_count_by_company_code[invoice.company_id] += 1  # invoice.company_id 是字串
    #     invoice_count_by_company_code[invoice.company.company_id] += 1
    # # 2️⃣ 建立公司代碼對應的 Company 資料（查主鍵）
    # company_map = {
    #     company.company_id: company for company in Company.objects.filter(company_id__in=invoice_count_by_company_code.keys())
        
    # }

    # # 3️⃣ 驗證每間公司是否有足夠的號碼可以使用
    # insufficient_companies = []

    # for company_code, count_needed in invoice_count_by_company_code.items():
    #     company_obj = company_map.get(company_code)
    #     if not company_obj:
    #         insufficient_companies.append(f"公司代碼 {company_code} 找不到對應公司資料")
    #         continue

    #     distributions = NumberDistribution.objects.filter(
    #         company=company_obj,
    #         status='available'
    #     )

    #     total_available = sum(
    #         int(d.end_number) - int(d.current_number or d.start_number) + 1
    #         for d in distributions
    #     )

    #     if total_available < count_needed:
    #         insufficient_companies.append(f"公司 {company_obj.company_name}（剩 {total_available} 張，需求 {count_needed} 張）")

    # if insufficient_companies:
    #     return HttpResponse("號碼不足，請檢查以下公司：\n" + "\n".join(insufficient_companies), status=400)

    # # 4️⃣ 準備號碼池（以 company.id 為 key）
    # number_pool = defaultdict(list)
    # for dist in NumberDistribution.objects.filter(status='available'):
    #     number_pool[dist.company.id].append(dist)

    # 5️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'B0101.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 6️⃣ 開始配號與寫入 Excel
    with transaction.atomic():
        for allowance in allowances:
            success_allowances = []
            validation_results = validate_allowance([allowance])
            validation_result = validation_results[0]
            is_valid_allowance = validation_result.get("is_allowance_valid", True)
            print(f"Allowance ID {allowance.id} valid? {is_valid_allowance}")
            if not is_valid_allowance:
                continue

            # #company_obj = company_map.get(invoice.company_id)
            # company_obj = company_map.get(invoice.company.company_id)
            # if not company_obj:
            #     #raise ValueError(f"找不到公司代碼為 {invoice.company_id} 的公司資料")
            #     raise ValueError(f"找不到公司代碼為 {invoice.company.company_id} 的公司資料")


            # distributions = sorted(
            #     number_pool[company_obj.id],
            #     key=lambda d: int(d.current_number or d.start_number)
            # )

            # 找一組號碼配給發票
            # assigned = False
            # for dist in distributions:
            #     current = int(dist.current_number or dist.start_number)
                # if current <= int(dist.end_number):
                # invoice_number = f"{dist.initial_char}{str(current).zfill(len(dist.start_number))}"
            allowance.allowance_status = '已開立'
            #allowance.allowance_date = localtime(timezone.now()).replace(tzinfo=None).date()
            #allowance.allowance_time = localtime(timezone.now()).replace(tzinfo=None).time()
            now = localtime(timezone.now()).replace(tzinfo=None)
            allowance.allowance_date= now.date() 
            allowance.allowance_time = now 
            allowance.export_date = now.date()
            allowance.save()
            success_allowances.append(allowance)
            
            for item in allowance.items.all():
                # 🟡 這裡是你要的邏輯：根據原始發票號碼與公司找發票，並更新其 allowance_status
                original_number = item.line_original_invoice_number
                if original_number:
                    try:
                        original_invoice = TWB2BMainItem.objects.filter(invoice_number=original_number).first()
                        if original_invoice.allowance_status != "已開立折讓單":
                            original_invoice.allowance_status = "已開立折讓單"
                            original_invoice.save()
                    except TWB2BMainItem.DoesNotExist:
                        pass  # 如找不到可記錄 log 或忽略
            
            config = CompanyNotificationConfig.objects.filter(company__company_id=allowance.company.company_id).first()

            output_dir_G0401 = config.output_dir_G0401
            output_dir_A4_paper = config.output_dir_allowance_A4_paper
            generate_G0401_xml_files(allowance, output_dir_G0401, G0401_XSD_path)
            generate_allowance_pdf(allowance, output_dir_A4_paper)

            for item in allowance.items.all():
                sheet.cell(row=row, column=1, value=allowance.company.company_id)
                sheet.cell(row=row, column=2, value=allowance.allowance_number)
                sheet.cell(row=row, column=3, value=allowance.allowance_date)
                sheet.cell(row=row, column=4, value=allowance.allowance_time)
                #sheet.cell(row=row, column=5, value=invoice.company.company_name)
                sheet.cell(row=row, column=5, value=allowance.allowance_type)
                sheet.cell(row=row, column=6, value=allowance.company.company_identifier)
                sheet.cell(row=row, column=7, value=allowance.buyer_identifier)
                sheet.cell(row=row, column=8, value=allowance.buyer_name)
                sheet.cell(row=row, column=9, value=item.line_sequence_number)
                sheet.cell(row=row, column=10, value=item.line_original_invoice_number)
                sheet.cell(row=row, column=11, value=item.line_original_invoice_date)
                sheet.cell(row=row, column=12, value=item.line_description)
                sheet.cell(row=row, column=13, value=item.line_quantity)
                sheet.cell(row=row, column=14, value=item.line_unit)
                sheet.cell(row=row, column=15, value=float(item.line_unit_price))
                sheet.cell(row=row, column=16, value=item.line_tax_type)
                sheet.cell(row=row, column=17, value=float(item.line_allowance_amount))
                sheet.cell(row=row, column=18, value=float(item.line_allowance_tax))
                sheet.cell(row=row, column=19, value=float(allowance.allowance_amount))
                sheet.cell(row=row, column=20, value=float(allowance.allowance_tax))
                row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="allowance.xlsx"'
    
    # 先找出所有選取的發票
    

    success_list = []
    excluded_list = []

    success_count = 0
    excluded_count = 0

    # 將 queryset 包進 list 傳入驗證函數
    validation_results = validate_allowance(allowances)

    for result in validation_results:
        allowance = result['allowance']
        allowance_number = getattr(allowance, 'allowance_number', f'ID {allowance.id}')

        if result['is_allowance_valid']:
            success_count += 1
            success_list.append(allowance_number)
        else:
            excluded_count += 1
            excluded_list.append({
                'allowance_number': allowance_number,
                'is_valid_amount': result['is_valid_amount'],
                'is_valid_tax': result['is_valid_tax'],
            })

    # 輸出結果（可改為寄信或 render template）
    print(f"✅ 成功數量: {success_count}")
    print(f"成功折讓單號: {', '.join(success_list)}")
    print(f"❌ 失敗數量: {excluded_count}")
    print("失敗明細：")
    for item in excluded_list:
        print(f"- 折讓單號: {item['allowance_number']}, 金額合法: {item['is_valid_amount']}, 稅額合法: {item['is_valid_tax']}")

    send_allowance_summary_email(to_email, success_count, excluded_count,success_list, excluded_list)
    messages.success(request, f"成功開立 {success_count} 張折讓單")
    
    redirect_url = reverse('twallowance')
    if return_querystring:
            redirect_url += '?' + return_querystring
    return redirect(redirect_url)


@csrf_exempt
def twallowance_delete_selected_invoices(request):
    config = SystemConfig.objects.first()
    to_email = config.operator_output_email_address if config else None
    F0401_XSD_path = config.F0401_XSD_path
    F0501_XSD_path = config.F0501_XSD_path
    G0401_XSD_path = config.G0401_XSD_path
    G0501_XSD_path = config.G0501_XSD_path
    A0201_XSD_path = config.A0201_XSD_path
    A0301_XSD_path = config.A0301_XSD_path
    A0101_XSD_path = config.A0101_XSD_path
    A0102_XSD_path = config.A0102_XSD_path
    A0202_XSD_path = config.A0202_XSD_path
    A0302_XSD_path = config.A0302_XSD_path
    B0101_XSD_path = config.B0101_XSD_path
    B0102_XSD_path = config.B0101_XSD_path
    B0201_XSD_path = config.B0201_XSD_path
    B0202_XSD_path = config.B0101_XSD_path
    if request.method == 'POST':
        selected_ids_raw = request.POST.get('selected_documents', '')
        selected_ids = selected_ids_raw.split(',') if selected_ids_raw else []
        
        return_querystring = request.POST.get('return_querystring', '')  # ⬅️ 關鍵步驟

        if not selected_ids:
            messages.error(request, "刪除失敗：未選擇任何發票。")
            if return_querystring:
                redirect_url += '?' + return_querystring
            return redirect(redirect_url)
        

        invalid_condition = Q(allowance_status='已開立') | Q(allowance_status='已作廢')
        allowance_to_delete = TWAllowance.objects.filter(id__in=selected_ids).prefetch_related('items').exclude(invalid_condition)
        deleted_count = allowance_to_delete.count()

        deleted_list=[]

        deleted_list = [
            f"{allowance.company.company_id} - {allowance.erp_number}"
            for allowance in allowance_to_delete
        ]

        allowance_to_delete.delete()

        if deleted_count > 0:
            send_allowance_deleted_email(to_email, deleted_count, deleted_list)

            messages.success(request, f"成功刪除 {deleted_count} 筆折讓單。")
        else:
            messages.warning(request, "所選取的發票皆為已開立發的折讓單未進行刪除。")
        
        redirect_url = reverse('twallowance')
        if return_querystring:
            redirect_url += '?' + return_querystring
        return redirect(redirect_url)
    else:
        return redirect('twallowance')
        
# @csrf_exempt
# def twallowance_update_void_status(request):
#     if request.method != 'POST':
#         return HttpResponse("Only POST allowed", status=405)

#     raw_ids = request.POST.get("selected_documents", "")
#     selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
#     if not selected_ids:
#         return HttpResponse("No invoice IDs provided", status=400)
    
#     allowances = TWAllowance.objects.filter(id__in=selected_ids).prefetch_related('items')
#     if not allowances.exists():
#         return HttpResponse("No invoices found", status=404)
   
   
#     #所有選取的發票皆為『未開立』發票或包含『已作廢』發票，無法作廢。
#     # 僅篩出「已開立」的折讓單
#     to_cancel = [a for a in allowances if a.allowance_status == '已開立']

#     if not to_cancel:
#         return HttpResponse("所有選取的折讓單皆為『未開立』或『已作廢』狀態，無法作廢。", status=400)

#     # ✅ 檢查是否每筆作廢發票都有填寫作廢理由
#     missing_reason = [allowance for allowance in to_cancel if not allowance.allowance_cancel_reason or allowance.allowance_cancel_reason.strip() == '']
#     if missing_reason:
#         return HttpResponse("有折讓單未填寫作廢理由，請補齊後再作廢。", status=400)
    

#     # 載入 Excel 樣板
#     template_path = os.path.join(settings.BASE_DIR, 'export', 'B0201.xlsx')
#     workbook = load_workbook(template_path)
#     sheet = workbook.active

#     row = 2  # Excel 開始列

#     with transaction.atomic():
#          for allowance in to_cancel:
#                         # 新增條件：只處理已開立的折讓單
#             if allowance.allowance_status != '已開立':
#                 # 可選擇跳過此筆或回傳錯誤訊息，這裡我先用 continue 跳過
#                 continue
#             # 更新作廢狀態與時間
#             allowance.allowance_status = '已作廢'
#             allowance.allowance_cancel_date = localtime(timezone.now()).replace(tzinfo=None).date()
#             allowance.allowance_cancel_time = localtime(timezone.now()).replace(tzinfo=None).time()
#             allowance.save()

#             output_dir_G0501 = r"C:\Users\waylin\mydjango\e_invoice\G0501"
#             xsd_path = r"C:\Users\waylin\mydjango\e_invoice\valid_xml\G0501.xsd"
#             generate_G0501_xml_files(allowance, output_dir_G0501, xsd_path)

#             # 生成 XML 檔案


#             sheet.cell(row=row, column=1, value=allowance.company.company_identifier)
#             sheet.cell(row=row, column=2, value=allowance.buyer_identifier)
#             sheet.cell(row=row, column=3, value=allowance.allowance_number)
#             sheet.cell(row=row, column=4, value=allowance.allowance_date)
#             sheet.cell(row=row, column=5, value=allowance.allowance_type)
#             sheet.cell(row=row, column=6, value=allowance.allowance_cancel_date)
#             sheet.cell(row=row, column=7, value=allowance.allowance_cancel_time)
#             sheet.cell(row=row, column=8, value=allowance.allowance_cancel_reason)
#             sheet.cell(row=row, column=9, value=allowance.allowance_cancel_remark)
#             row += 1

#     # 匯出 Excel
#     output = BytesIO()
#     workbook.save(output)
#     output.seek(0)

#     response = HttpResponse(
#         output,
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="B0201.xlsx"'
#     return response

@csrf_exempt
def twallowance_update_cancel_status(request):
    config = SystemConfig.objects.first()
    to_email = config.operator_output_email_address if config else None
    F0401_XSD_path = config.F0401_XSD_path
    F0501_XSD_path = config.F0501_XSD_path
    G0401_XSD_path = config.G0401_XSD_path
    G0501_XSD_path = config.G0501_XSD_path
    A0201_XSD_path = config.A0201_XSD_path
    A0301_XSD_path = config.A0301_XSD_path
    A0101_XSD_path = config.A0101_XSD_path
    A0102_XSD_path = config.A0102_XSD_path
    A0202_XSD_path = config.A0202_XSD_path
    A0302_XSD_path = config.A0302_XSD_path
    B0101_XSD_path = config.B0101_XSD_path
    B0102_XSD_path = config.B0101_XSD_path
    B0201_XSD_path = config.B0201_XSD_path
    B0202_XSD_path = config.B0101_XSD_path
    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        allowance_id = data.get("allowance_id")
        cancel_reason = data.get("allowance_cancel_reason", "").strip()
        cancel_remark = data.get("allowance_cancel_remark", "").strip()

        if not allowance_id:
            return JsonResponse({"success": False, "message": "缺少折讓單 ID"}, status=400)
        if not cancel_reason:
            return JsonResponse({"success": False, "message": "請輸入作廢理由"}, status=400)
        try:
            allowance= TWAllowance.objects.select_related('company').get(id=allowance_id)
        except TWAllowance.DoesNotExist:
            return JsonResponse({"success": False, "message": "找不到該折讓單"}, status=404)

        if allowance.allowance_status == '未開立':
            return JsonResponse({"success": False, "message": "該折讓單為未開立狀態，無法作廢"}, status=400)
        elif allowance.allowance_status == '已作廢':
            return JsonResponse({"success": False, "message": "該折讓單已作廢，無法再次作廢"}, status=400)
        #elif allowance.allowance_status == '已開立折讓單':
        #    return JsonResponse({"success": False, "message": "該折讓單已開立折讓單，無法作廢"}, status=400)
        #    messages.error(request, f"該折讓單已開立折讓單，無法作廢")
        elif allowance.mof_response != 'S0001':
            return JsonResponse({"success": False, "message": "該折讓單稅局未認證，無法作廢"}, status=400)

        
        
        now = localtime()

        allowance.allowance_cancel_date = now.date()
        allowance.allowance_cancel_time = now.time()
        allowance.allowance_cancel_reason = cancel_reason
        allowance.allowance_cancel_remark = cancel_remark
        allowance.allowance_status = "已作廢"
        allowance.save()

        # current_roc_year = now.year - 1911
        # current_roc_month = now.month
        # current_period = current_roc_year * 100 + current_roc_month  # e.g., 11502
        
        # ✅ 檢查是否跨期作廢 → 要填 returntax_document_number
        # try:
        #     invoice_period = int(invoice.invoice_period)  # 確保是整數，如 11412
        # except (ValueError, TypeError):
        #     return JsonResponse({"success": False, "message": "發票期別資料異常"}, status=400)
        
        # if invoice_period < current_period:
        #     if not invoice.returntax_document_number or invoice.returntax_document_number.strip() == '':
        #         return JsonResponse({
        #             "success": False,
        #             "message": "跨期作廢的發票需填寫折讓參考號（returntax_document_number）。"
        #         }, status=400)

        # 產生 G0501 XML
        output_dir_F0501 = r"C:\Users\waylin\mydjango\e_invoice\G0501"
        xsd_path = r"C:\Users\waylin\mydjango\e_invoice\valid_xml\G0501.xsd"
        generate_G0501_xml_files(allowance, output_dir_F0501, xsd_path)

        success_list = [f"{allowance.company.company_name} - {allowance.allowance_number}"]
        excluded_list = []
        send_allowance_canceled_email(to_email, success_count=1, excluded_count=0, success_list=success_list, excluded_list=excluded_list)

        allowance.save()
        return JsonResponse({"success": True, "message": "作廢成功並產出 G0501.xml"})
 
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "資料格式錯誤"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
#======================================================驗證發票明細=======================================================


