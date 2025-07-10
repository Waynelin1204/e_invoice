# ====== Python 標準函式庫 ======
import os
import json
import logging
import subprocess
from io import BytesIO
from datetime import datetime, timedelta
from collections import defaultdict
import xml.etree.ElementTree as ET
from django.utils.timezone import localtime
from decimal import Decimal, InvalidOperation
import random


# ====== 第三方套件 ======
import pandas as pd
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# ====== Django 基礎功能 ======
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.utils import timezone

# ====== Django DB 操作 ======
from django.db import connection
from django.db import transaction
from django.db.models import Q, Count, Sum

# ====== Django 使用者與驗證 ======
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test

# ====== 專案內部（Model 與 Form） ======
from e_invoices.models import (
    RegisterForm, LoginForm,
    Twa0101, Twa0101Item, Ocr, Ocritem, Company, UserProfile,
    NumberDistribution, TWB2BMainItem, TWB2BLineItem, TWAllowanceLineItem, TWAllowance
)
from e_invoices.forms import NumberDistributionForm
from e_invoices.services import generate_invoice_B2C_a4, generate_invoice_B2B_a4, generate_F0401_xml_files,generate_F0701_xml_files,generate_invoice_B2B_format25_pdf,  generate_F0501_xml_files
from e_invoices.services import send_invoice_summary_email, generate_invoice_B2B_format25_pdf_stamp, send_invoice_canceled_email

@login_required
def twb2bmainitem(request):
    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile
    
    # 取得該使用者可查看的公司名稱列表
    #viewable_company_codes = user_profile.viewable_companies.values_list('id', flat=True)
    viewable_company_codes = user_profile.viewable_companies.values_list('company_id', flat=True)

    
    # 計算從今天開始往前推的60天的日期
    sixty_days_ago = datetime.today() - timedelta(days=60)
    
    # 篩選條件：只顯示最近60天的發票
    filter_conditions = Q(company__in=viewable_company_codes) & Q(erp_date__gte=sixty_days_ago)

    # 查詢符合條件的資料，並使用 prefetch_related 來查詢發票明細
    documents = TWB2BMainItem.objects.filter(filter_conditions).prefetch_related('items').order_by('-erp_date')
    company_options = user_profile.viewable_companies.all()



#-----------------計算剩餘可扣抵金額-----------------
    # for document in documents:
    #     document.is_disabled = (
    #         document.invoice_status == '已作廢' or
    #         (
    #             document.tax_type == '2' and (
    #                 (document.zerotax_sales_amount or 0) <= 0 or
    #                 not document.customs_clearance_mark or
    #                 not document.bonded_area_confirm or
    #                 not document.zero_tax_rate_reason
    #             )
    #         )
    #     )
    # # 批量更新邏輯
    # for document in documents:
    #     total_amount = document.total_amount or 0
    #     accurated_allowance_amount = document.accurated_allowance_amount or 0
    #     document.remaining_allowance_amount = total_amount - accurated_allowance_amount

    # # 使用 bulk_update 批量保存
    # TWB2BMainItem.objects.bulk_update(documents, ['remaining_allowance_amount'])
#-------------------------------------------------------
# 


    # 分頁：每頁顯示25筆資料
    paginator = Paginator(documents, 25)  # 每頁顯示25筆資料
    page_number = request.GET.get('page')  # 取得當前頁數
    page_obj = paginator.get_page(page_number)  # 根據頁數取得對應的資料

    # 傳遞資料給模板
    context = {
        'documents': page_obj,  # 傳遞分頁後的資料
        "company_options": company_options,
    }
    print("🔍 可查看的公司 company_id：", list(viewable_company_codes))
    print("✅ 撈到的發票數：", TWB2BMainItem.objects.filter(filter_conditions).count())
    print("🟡 viewable_company_ids:", list(viewable_company_codes))
    print("🟡 篩選時間從:", sixty_days_ago)
    #print("🟡 所有發票的公司代碼：", TWB2BMainItem.objects.values_list("company_code", flat=True).distinct())
    print("🟡 所有發票的公司代碼：", TWB2BMainItem.objects.values_list("company_id", flat=True).distinct())    
    print("🟡 最近60天的發票：", TWB2BMainItem.objects.filter(erp_date__gte=sixty_days_ago).values_list("company_id", flat=True))
    print("🟡 完整符合條件的發票數：", TWB2BMainItem.objects.filter(filter_conditions).count())

    return render(request, 'twb2bmainitem.html', context)

#======================================================B2B發票明細=======================================================
    
def twb2blineitem(request, id):
    document = get_object_or_404(TWB2BMainItem, id=id)
    items = document.items.all()  # 確保有正確查詢
    # 計算屬性值
    # total_allowanced_amount = document.allowance_lineitems.aggregate(
    #     total=Sum('line_allowance_amount')
    # )['total'] or 0

    # total_allowanced_tax = document.allowance_lineitems.aggregate(
    #     total=Sum('line_allowance_tax')
    # )['total'] or 0

# 對應到 linked_invoice（ForeignKey）所連的 TWB2BMainItem 的 invoice_number 欄位
    total_allowanced_amount = TWAllowanceLineItem.objects.filter(
        linked_invoice=document,
        twallowance__allowance_status='已開立'
    ).aggregate(
        total_allowanced_amount=Sum('line_allowance_amount')
    )['total_allowanced_amount'] or 0

    total_allowanced_tax = TWAllowanceLineItem.objects.filter(
        linked_invoice=document,
        twallowance__allowance_status='已開立'
    ).aggregate(
        total_allowanced_tax=Sum('line_allowance_tax')
    )['total_allowanced_tax'] or 0

    # 發票可用金額與稅額
    available_amount = (
        (document.sales_amount or 0)
        + (document.zerotax_sales_amount or 0)
        + (document.freetax_sales_amount or 0)
        - total_allowanced_amount
    )

    available_tax = (document.total_tax_amount or 0) - total_allowanced_tax
    
    related_allowances = TWAllowance.objects.filter(
        items__linked_invoice=document,
        allowance_status='已開立'
    ).distinct()




    # 判斷是否有效
    print(f"原銷售金額: {document.sales_amount}")    
    print(f"原零稅銷售金額: {document.zerotax_sales_amount}")
    print(f"原免稅銷售金額: {document.freetax_sales_amount}")
    print(f"已開立折讓單號: {[a.allowance_number for a in related_allowances]}")
    print(f"累計折讓金額: {total_allowanced_amount}")
    print(f"累計折讓金額: {total_allowanced_tax}")
    print(f"剩餘折讓金額: {available_amount}")
    print(f"剩餘折讓稅額: {available_tax}")



    return render(request, 'document/twb2blineitem.html', {
        'document': document,
        'items': items,
        'available_amount': available_amount,
        'available_tax': available_tax,
        'total_allowanced_amount': total_allowanced_amount,
        'total_allowanced_tax': total_allowanced_tax,
        'related_allowances': related_allowances,
    })

@csrf_exempt
def twb2blineitem_update(request, id):
    document = get_object_or_404(TWB2BMainItem, id=id)
    if request.method == 'POST':
        # 處理主項目資料
        #invoice_period = request.POST.get('invoice_period', '').strip()
        erp_reference = request.POST.get('erp_reference', '').strip()
        seller_bp_id = request.POST.get('seller_bp_id', '').strip()
        buyer_bp_id = request.POST.get('buyer_bp_id', '').strip()
        buyer_remark = request.POST.get('buyer_remark', '').strip()
        main_remark = request.POST.get('main_remark', '').strip()
        customs_clearance_mark = request.POST.get('customs_clearance_mark', '').strip()
        category = request.POST.get('category', '').strip()
        relate_number = request.POST.get('relate_number', '').strip()
        bonded_area_confirm = request.POST.get('bonded_area_confirm', '').strip()
        zero_tax_rate_reason = request.POST.get('zero_tax_rate_reason', '').strip()
        reserved1 = request.POST.get('reserved1', '').strip()
        reserved2 = request.POST.get('reserved2', '').strip()
        cancel_reason = request.POST.get('cancel_reason', '').strip()
        returntax_document_number = request.POST.get('returntax_document_number', '').strip()
        try:
            original_currency_amount = Decimal(request.POST.get('original_currency_amount', '').strip()) if request.POST.get('original_currency_amount', '').strip() else None
        except InvalidOperation:
            original_currency_amount = None
        try:
            exchange_rate = Decimal(request.POST.get('exchange_rate', '').strip()) if request.POST.get('exchange_rate', '').strip() else None
        except InvalidOperation:
            exchange_rate = None
        currency = request.POST.get('currency', '').strip()

        # 更新主項目資料
        #document.invoice_period = invoice_period
        document.erp_reference = erp_reference
        document.seller_bp_id = seller_bp_id
        document.buyer_bp_id = buyer_bp_id
        document.buyer_remark = buyer_remark
        document.main_remark = main_remark
        document.customs_clearance_mark = customs_clearance_mark
        document.category = category
        document.relate_number = relate_number
        document.bonded_area_confirm = bonded_area_confirm
        document.zero_tax_rate_reason = zero_tax_rate_reason
        document.reserved1 = reserved1
        document.reserved2 = reserved2
        document.original_currency_amount = original_currency_amount
        document.exchange_rate = exchange_rate
        document.currency = currency
        document.cancel_reason= cancel_reason
        document.returntax_document_number = returntax_document_number
        document.save()  # 保存主項目資料

        # 更新明細項目資料
        for item in document.items.all():
            line_remark = request.POST.get(f'line_remark_{item.id}', '').strip()
            line_relate_number = request.POST.get(f'line_relate_number_{item.id}', '').strip()

            item.line_remark = line_remark
            item.line_relate_number = line_relate_number
            item.save()  # 保存明細項目資料

        return redirect('twb2blineitem', id=document.id)  # 更新後重定向到發票詳情頁面

    return render(request, 'document/twb2blineitem.html', {'document': document})


#======================================================B2B發票篩選=======================================================

def twb2bmainitem_filter(request):
    # 預設顯示筆數
    display_limit = int(request.GET.get("display_limit", 20))  # 默認為20筆資料

    # 其他篩選條件
    invoice_status_filter = request.GET.get("invoice_status")
    #void_status_filter = request.GET.get("void_status")
    tax_type_filter = request.GET.get("tax_type")
    company_id_filter = request.GET.get("company_id")  # 新增公司篩選條件


    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile

    # 取得該使用者可查看的公司名稱列表
    company_options = user_profile.viewable_companies.all()
    #viewable_company_codes = user_profile.viewable_companies.values_list('id', flat=True)
    viewable_company_codes = user_profile.viewable_companies.values_list('company_id', flat=True)



    # 計算兩個月前的日期
    two_months_ago = datetime.today() - timedelta(days=60)

    # 取得時間範圍的過濾條件（默認從兩個月前開始）
    start_date = request.GET.get("start_date", two_months_ago.strftime('%Y-%m-%d'))
    end_date = request.GET.get("end_date", datetime.today().strftime('%Y-%m-%d'))

    # 轉換成 datetime 格式
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    except ValueError:
        start_date = two_months_ago  # 如果格式錯誤，使用預設的兩個月前日期

    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        end_date = datetime.today()  # 如果格式錯誤，使用今天的日期

    if (end_date - start_date).days > 60:
        messages.error(request, "查詢區間不得超過 60 天")
        return redirect('twb2bmainitem')
    if start_date > end_date:
        messages.error(request, "開始日期不能大於結束日期")
        return redirect('twb2bmainitem')
    
    # 查詢條件構建
    filters = Q()
    if invoice_status_filter:
        filters &= Q(invoice_status=invoice_status_filter)
    
    #if void_status_filter:
    #    filters &= Q(void_status=void_status_filter)
    if tax_type_filter:
        filters &= Q(tax_type=tax_type_filter)
    if company_id_filter:
        filters &= Q(company_id=company_id_filter)

    # 限制在兩個月前到今天的時間範圍內
    filters &= Q(erp_date__range=[start_date, end_date])


    # 加入公司權限過濾條件
    filters &= Q(company__in=viewable_company_codes)

    # 查詢所有符合條件的發票資料
    documents = TWB2BMainItem.objects.filter(filters).order_by('-erp_date')

    for document in documents:
        document.is_disabled = (
            document.invoice_status == '已作廢' or
            (
                document.tax_type == '2' and (
                    (document.zerotax_sales_amount or 0) <= 0 or
                    not document.customs_clearance_mark or
                    not document.bonded_area_confirm or
                    not document.zero_tax_rate_reason
                )
            )
        )


    # 分頁
    paginator = Paginator(documents, display_limit)  # 每頁顯示的資料筆數
    page_number = request.GET.get('page')  # 獲取當前頁碼
    page_obj = paginator.get_page(page_number)  # 根據頁碼獲取相應的頁面資料

    # 獲取篩選條件的選項
    invoice_status = TWB2BMainItem.objects.values_list('invoice_status', flat=True).distinct()
    #void_status = TWB2BMainItem.objects.values_list('void_status', flat=True).distinct()
    tax_type = TWB2BMainItem.objects.values_list('tax_type', flat=True).distinct()

    # 檢查公司ID篩選是否有效
    if company_id_filter and company_id_filter not in viewable_company_codes:
        messages.error(request, "您無權限查看該公司資料")
        return redirect('twb2bmainitem')
    

    return render(request, "twb2bmainitem_filter.html", {
        "company_id_filter": company_id_filter,
        "documents": page_obj,  # 傳遞分頁結果
        "company_options": company_options,
        "invoice_status": invoice_status,
        "tax_type_filter": tax_type_filter,
        "invoice_status_filter": invoice_status_filter,
        #"void_status": void_status,
        "tax_type": tax_type,
        "display_limit": display_limit,  # 傳遞選擇的筆數
        "start_date": start_date.strftime('%Y-%m-%d'),  # 顯示篩選的開始日期
        "end_date": end_date.strftime('%Y-%m-%d'),  # 顯示篩選的結束日期
    })

#======================================================B2B發票匯出=======================================================

aes_key = "73BE69AA9826613AFAEC11C215F08302"
output_path = r"C:\\Users\\waylin\\mydjango\\e_invoice\\print\\"

@csrf_exempt
def twb2bmainitem_export_invoices(request):
    if request.method != 'POST':
        return HttpResponse("Only POST allowed", status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    if not selected_ids:
        return HttpResponse("No invoice IDs provided", status=400)

    #invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items')
    # 零稅率欄位不合法的條件
    zero_tax_invalid = (
        Q(tax_type='2') & (
            Q(zerotax_sales_amount__lte=0) |
            Q(customs_clearance_mark__in=[None, '']) |
            Q(bonded_area_confirm__in=[None, '']) |
            Q(zero_tax_rate_reason__in=[None, ''])
        )
    )

    # 綜合排除條件
    invalid_condition = Q(invoice_status='已開立') | zero_tax_invalid

    invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items').exclude(invalid_condition)

    if not invoices.exists():
         return HttpResponse("No invoices found", status=404)

    # 先找出所有選取的發票
    #all_selected_invoices = TWB2BMainItem.objects.filter(id__in=selected_ids)

    # 計算「已開立」的數量
    #excluded_count = all_selected_invoices.filter(invoice_status='已開立').count()
    #excluded_count = invoices.count()

    # 篩選出尚未開立的發票
    #invoices = all_selected_invoices.exclude(invoice_status='已開立').prefetch_related('items')
    


    # 1️⃣ 統計各公司所需發票數
    #invoice_count_by_company_code = defaultdict(int)
    invoice_count_by_company_id = defaultdict(int)

    for invoice in invoices:
        #invoice_count_by_company_code[invoice.company_id] += 1  # invoice.company_id 是字串
        #invoice_count_by_company_code[invoice.company.company_id] += 1
        invoice_count_by_company_id[invoice.company.company_id] += 1

    # 2️⃣ 建立公司代碼對應的 Company 資料（查主鍵）
    company_map = {
        #company.company_id: company for company in Company.objects.filter(company_id__in=invoice_count_by_company_code.keys())
        company.company_id: company for company in Company.objects.filter(company_id__in=invoice_count_by_company_id.keys())
    
    }

    # 3️⃣ 驗證每間公司是否有足夠的號碼可以使用
    insufficient_companies = []

    #for company_code, count_needed in invoice_count_by_company_code.items():
    for company_id, count_needed in invoice_count_by_company_id.items():
        #company_obj = company_map.get(company_code)
        company_obj = company_map.get(company_id)
        if not company_obj:
            #insufficient_companies.append(f"公司代碼 {company_code} 找不到對應公司資料")
            insufficient_companies.append(f"公司代碼 {company_id} 找不到對應公司資料")
            continue

        distributions = NumberDistribution.objects.filter(
            company=company_obj,
            status='available'
        )

        total_available = sum(
            int(d.end_number) - int(d.current_number or d.start_number) + 1
            for d in distributions
        )

        if total_available < count_needed:
            insufficient_companies.append(f"公司 {company_obj.company_name}（剩 {total_available} 張，需求 {count_needed} 張）")

    if insufficient_companies:
        return HttpResponse("號碼不足，請檢查以下公司：\n" + "\n".join(insufficient_companies), status=400)

    # 4️⃣ 準備號碼池（以 company.id 為 key）
    number_pool = defaultdict(list)
    for dist in NumberDistribution.objects.filter(status='available'):
        #number_pool[dist.company.id].append(dist)
        number_pool[dist.company.company_id].append(dist)

    # 5️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'A0101.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 6️⃣ 開始配號與寫入 Excel
    with transaction.atomic():
        for invoice in invoices:
            random_codes = str(random.randint(0, 9999)).zfill(4) # 4碼隨機碼
            #company_obj = company_map.get(invoice.company_id)
            company_obj = company_map.get(invoice.company.company_id)
            if not company_obj:
                #raise ValueError(f"找不到公司代碼為 {invoice.company_id} 的公司資料")
                raise ValueError(f"找不到公司代碼為 {invoice.company.company_id} 的公司資料")


            distributions = sorted(
                #number_pool[company_obj.id],
                number_pool[company_obj.company_id],
                key=lambda d: int(d.current_number or d.start_number)
            )

            success_invoices = []

            # 找一組號碼配給發票
            assigned = False
            for dist in distributions:
                current = int(dist.current_number or dist.start_number)
                invoice_number = f"{dist.initial_char}{str(current).zfill(len(dist.start_number))}"

                if current <= int(dist.end_number):
                    if invoice.invoice_number:
                        invoice.invoice_number = invoice.invoice_number
                        invoice.invoice_date = invoice.invoice_date
                        invoice.invoice_time = invoice.invoice_time
                        invoice.invoice_period = invoice.invoice_period
                        invoice.random_code = invoice.random_code
                        invoice.invoice_status = '註銷後重開'
                    else:    
                        invoice.invoice_number = invoice_number
                        invoice.invoice_date = localtime(timezone.now()).replace(tzinfo=None).date()
                        invoice.invoice_time = localtime(timezone.now()).strftime('%H:%M:%S')
                        invoice.invoice_period = dist.period
                        invoice.random_code = random_codes
                        dist.current_number = str(current + 1).zfill(len(dist.start_number))
                        dist.last_used_date = timezone.now().date()
                        dist.save()
                        invoice.invoice_status = '已開立'

                    
                    invoice.export_date = localtime(timezone.now()).replace(tzinfo=None).date()
                    
                    invoice.save()
                    success_invoices.append(invoice)
                    assigned = True
                    break

            if not assigned:
                raise ValueError(f"公司 {company_obj.company_name} 號碼區間不足")
            
            #img_path = generate_invoice_image_qrcodes(invoice, aes_key)
            #output_path = r"C:\Users\waylin\mydjango\e_invoice\print"

            #output_dir = r"C:\Users\waylin\mydjango\e_invoice\print"
            output_dir_F0401 = r"C:\Users\waylin\mydjango\e_invoice\F0401"
            xsd_path = r"C:\Users\waylin\mydjango\e_invoice\valid_xml\F0401.xsd"
            

            #img_path = generate_invoice_a4(invoice, aes_key, output_dir)


            generate_invoice_B2C_a4(invoice, aes_key, output_path, random_codes)
            generate_invoice_B2B_a4(invoice, aes_key, output_path)
            generate_F0401_xml_files(invoice, output_dir_F0401, xsd_path,random_codes)
            generate_invoice_B2B_format25_pdf(invoice, output_path)
            generate_invoice_B2B_format25_pdf_stamp(invoice, output_path)


            # 寫入 Excel
            for item in invoice.items.all():
                sheet.cell(row=row, column=1, value=invoice.company.company_id)
                sheet.cell(row=row, column=2, value=invoice.invoice_number)
                sheet.cell(row=row, column=3, value=invoice.invoice_date)
                sheet.cell(row=row, column=4, value=invoice.invoice_time)
                #sheet.cell(row=row, column=5, value=invoice.company.company_name)
                sheet.cell(row=row, column=5, value=invoice.invoice_type)
                sheet.cell(row=row, column=6, value=invoice.company.company_identifier)
                sheet.cell(row=row, column=7, value=invoice.buyer_identifier)
                sheet.cell(row=row, column=8, value=invoice.buyer_name)
                sheet.cell(row=row, column=9, value=invoice.buyer_bp_id)
                sheet.cell(row=row, column=10, value=invoice.buyer_remark)
                sheet.cell(row=row, column=11, value=invoice.main_remark)
                sheet.cell(row=row, column=12, value=invoice.customs_clearance_mark)
                sheet.cell(row=row, column=13, value=invoice.category)
                sheet.cell(row=row, column=14, value=invoice.relate_number)
                sheet.cell(row=row, column=15, value=invoice.bonded_area_confirm)
                sheet.cell(row=row, column=16, value=invoice.zero_tax_rate_reason)
                sheet.cell(row=row, column=17, value=invoice.reserved1)
                sheet.cell(row=row, column=18, value=invoice.reserved2)
                sheet.cell(row=row, column=19, value=invoice.sales_amount)
                sheet.cell(row=row, column=20, value=invoice.freetax_sales_amount)
                sheet.cell(row=row, column=21, value=invoice.zerotax_sales_amount)
                sheet.cell(row=row, column=22, value=invoice.tax_type)
                sheet.cell(row=row, column=23, value=invoice.tax_rate)
                sheet.cell(row=row, column=24, value=float(invoice.total_tax_amount))
                sheet.cell(row=row, column=25, value=float(invoice.total_amount))
                sheet.cell(row=row, column=26, value=invoice.original_currency_amount)
                sheet.cell(row=row, column=27, value=invoice.exchange_rate)
                sheet.cell(row=row, column=28, value=invoice.currency)
                sheet.cell(row=row, column=29, value=item.line_description)
                sheet.cell(row=row, column=30, value=item.line_quantity)
                sheet.cell(row=row, column=31, value=item.line_unit)
                sheet.cell(row=row, column=32, value=float(item.line_unit_price))
                sheet.cell(row=row, column=33, value=item.line_tax_type)
                sheet.cell(row=row, column=34, value=float(item.line_amount))
                sheet.cell(row=row, column=35, value=item.line_sequence_number)
                sheet.cell(row=row, column=36, value=item.line_remark)
                sheet.cell(row=row, column=37, value=item.line_relate_number)
                row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'

    # 先找出所有選取的發票
    all_selected_invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items')
    
    # 被排除發票
    
    excluded_invoices = []

    for inv in all_selected_invoices:
        reasons = []
        if inv.tax_type == '2':
            if inv.zerotax_sales_amount <= 0:
                reasons.append("零稅率金額為 0")
            if not inv.customs_clearance_mark:
                reasons.append("缺少報關註記")
            if not inv.bonded_area_confirm:
                reasons.append("缺少保稅註記")
            if not inv.zero_tax_rate_reason:
                reasons.append("缺少零稅率原因")

        if reasons:
            excluded_invoices.append({
                "company_name": inv.company.company_name,
                "invoice_number": inv.invoice_number or "(尚未配號)",
                "reasons": reasons
            })

        # 這裡才正確
        success_count = len(success_invoices)
        excluded_count = len(excluded_invoices)

        success_list = [
            f"{inv.company.company_name} - {inv.invoice_number}"
            for inv in success_invoices
        ]
        excluded_list = [
            f"{e['company_name']} - {e['invoice_number']}：{', '.join(e['reasons'])}"
            for e in excluded_invoices
        ]
        to_email = "waylin@deloitte.com.tw"

    send_invoice_summary_email(to_email, success_count, excluded_count,success_list, excluded_list)
    
    return response

@csrf_exempt
def twb2bmainitem_export_invoices_wo_number(request):
    if request.method != 'POST':
        return HttpResponse("Only POST allowed", status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    if not selected_ids:
        return HttpResponse("No invoice IDs provided", status=400)

    # 零稅率欄位不合法的條件
    zero_tax_invalid = (
        Q(tax_type='2') & (
            Q(zerotax_sales_amount__lte=0) |
            Q(customs_clearance_mark__in=[None, '']) |
            Q(bonded_area_confirm__in=[None, '']) |
            Q(zero_tax_rate_reason__in=[None, ''])
        )
    )

    # 綜合排除條件
    invalid_condition = Q(invoice_status='已開立') | zero_tax_invalid

    invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items').exclude(invalid_condition)

    if not invoices.exists():
        return HttpResponse("No invoices found", status=404)
    



    # 1️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'A0101.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 2️⃣ 寫入 Excel
    for invoice in invoices:
        invoice.invoice_status = '已開立'
        invoice.invoice_date = localtime(timezone.now()).replace(tzinfo=None).date()
        invoice.export_date = localtime(timezone.now()).replace(tzinfo=None).date()
        invoice.save()
        for item in invoice.items.all():
            sheet.cell(row=row, column=1, value=invoice.company.company_id)
            sheet.cell(row=row, column=2, value=invoice.invoice_number)
            sheet.cell(row=row, column=3, value=invoice.invoice_date)
            sheet.cell(row=row, column=4, value=invoice.invoice_time)
            sheet.cell(row=row, column=5, value=invoice.invoice_type)
            sheet.cell(row=row, column=6, value=invoice.company.company_identifier)
            sheet.cell(row=row, column=7, value=invoice.buyer_identifier)
            sheet.cell(row=row, column=8, value=invoice.buyer_name)
            sheet.cell(row=row, column=9, value=invoice.buyer_bp_id)
            sheet.cell(row=row, column=10, value=invoice.buyer_remark)
            sheet.cell(row=row, column=11, value=invoice.main_remark)
            sheet.cell(row=row, column=12, value=invoice.customs_clearance_mark)
            sheet.cell(row=row, column=13, value=invoice.category)
            sheet.cell(row=row, column=14, value=invoice.relate_number)
            sheet.cell(row=row, column=15, value=invoice.bonded_area_confirm)
            sheet.cell(row=row, column=16, value=invoice.zero_tax_rate_reason)
            sheet.cell(row=row, column=17, value=invoice.reserved1)
            sheet.cell(row=row, column=18, value=invoice.reserved2)
            sheet.cell(row=row, column=19, value=invoice.sales_amount)
            sheet.cell(row=row, column=20, value=invoice.freetax_sales_amount)
            sheet.cell(row=row, column=21, value=invoice.zerotax_sales_amount)
            sheet.cell(row=row, column=22, value=invoice.tax_type)
            sheet.cell(row=row, column=23, value=invoice.tax_rate)
            sheet.cell(row=row, column=24, value=float(invoice.total_tax_amount))
            sheet.cell(row=row, column=25, value=float(invoice.total_amount))
            sheet.cell(row=row, column=26, value=invoice.original_currency_amount)
            sheet.cell(row=row, column=27, value=invoice.exchange_rate)
            sheet.cell(row=row, column=28, value=invoice.currency)
            sheet.cell(row=row, column=29, value=item.line_description)
            sheet.cell(row=row, column=30, value=item.line_quantity)
            sheet.cell(row=row, column=31, value=item.line_unit)
            sheet.cell(row=row, column=32, value=float(item.line_unit_price))
            sheet.cell(row=row, column=33, value=item.line_tax_type)
            sheet.cell(row=row, column=34, value=float(item.line_amount))
            sheet.cell(row=row, column=35, value=item.line_sequence_number)
            sheet.cell(row=row, column=36, value=item.line_remark)
            sheet.cell(row=row, column=37, value=item.line_relate_number)
            row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
    return response

@csrf_exempt
def twb2bmainitem_delete_selected_invoices(request):
    if request.method == 'POST':
        selected_ids_raw = request.POST.get('selected_documents', '')
        selected_ids = selected_ids_raw.split(',') if selected_ids_raw else []

        if not selected_ids:
            messages.error(request, "刪除失敗：未選擇任何發票。")
            return redirect('twb2bmain')

        # 只刪除 invoice_status 為「未開立」的發票
        invoices_to_delete = TWB2BMainItem.objects.filter(id__in=selected_ids, invoice_status="未開立")
        deleted_count, _ = invoices_to_delete.delete()

        if deleted_count > 0:
            messages.success(request, f"成功刪除 {deleted_count} 筆發票。")
        else:
            messages.warning(request, "未找到對應的發票資料，未進行刪除。")

        return redirect('twb2bmainitem')
    else:
        return redirect('twb2bmainitem')
        
# @csrf_exempt
# def twb2bmainitem_update_void_status(request):
#     if request.method != 'POST':
#         return HttpResponse("Only POST allowed", status=405)

#     raw_ids = request.POST.get("selected_documents", "")
#     selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
#     if not selected_ids:
#         return HttpResponse("No invoice IDs provided", status=400)

#     invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items')
#     if not invoices.exists():
#         return HttpResponse("No invoices found", status=404)


#     # ✅ 排除未開立的發票
#     #not_issued = invoices.filter(invoice_status='未開立')
#     to_cancel = invoices.exclude(Q(invoice_status='未開立') | Q(allowance_status='已開立折讓單'))


#     if not to_cancel.exists():
#         return HttpResponse("所有選取的發票皆為『未開立』發票或包含『已開立折讓單』發票，故無法作廢。", status=400)
    
#     # ✅ 檢查是否每筆作廢發票都有填寫作廢理由
#     missing_reason = [inv for inv in to_cancel if not inv.cancel_reason or inv.cancel_reason.strip() == '']
#     if missing_reason:
#         return HttpResponse("有發票未填寫作廢理由，請補齊後再作廢。", status=400)
    
#     now = datetime.now()
#     current_roc_year = now.year - 1911
#     current_roc_month = now.month
#     current_period = current_roc_year * 100 + current_roc_month  # e.g., 11502
    
#     # ✅ 檢查是否跨期作廢 → 要填 returntax_document_number
#     cross_period_missing = []
#     for invoices in to_cancel:
#         try:
#             invoice_period = int(invoices.invoice_period)  # 確保是整數，如 11412
#         except (ValueError, TypeError):
#             continue  # 如果 invoice_period 不正確就跳過，視情況也可視為錯誤

#         if invoice_period < current_period:
#             if not invoices.returntax_document_number or invoices.returntax_document_number.strip() == '':
#                 cross_period_missing.append(invoices)

#     if cross_period_missing:
#         return HttpResponse("跨期作廢的發票需填寫折讓參考號（returntax_document_number）。", status=400)

#     output_dir_F0501 = r"C:\Users\waylin\mydjango\e_invoice\F0501"
#     xsd_path = r"C:\Users\waylin\mydjango\e_invoice\valid_xml\F0501.xsd"

    
#     # 載入 Excel 樣板
#     template_path = os.path.join(settings.BASE_DIR, 'export', 'F0501.xlsx')
#     workbook = load_workbook(template_path)
#     sheet = workbook.active

#     row = 2  # Excel 開始列

#     with transaction.atomic():
#         for invoice in to_cancel:
#             # 更新作廢狀態與時間
#             invoice.invoice_status = '已作廢'
#             invoice.cancel_date = localtime(timezone.now()).replace(tzinfo=None).date()
#             invoice.cancel_time = localtime(timezone.now()).replace(tzinfo=None).time()
#             invoice.original_invoice_date = invoice.invoice_date
#             invoice.original_invoice_number = invoice.invoice_number
#             invoice.save()

#             generate_F0501_xml_files(invoices, output_dir_F0501, xsd_path)


#             sheet.cell(row=row, column=1, value=invoice.company.company_identifier)
#             sheet.cell(row=row, column=2, value=invoice.buyer_identifier)
#             sheet.cell(row=row, column=3, value=invoice.invoice_date)
#             sheet.cell(row=row, column=4, value=invoice.invoice_number)
#             sheet.cell(row=row, column=5, value=invoice.invoice_period)
#             sheet.cell(row=row, column=6, value=invoice.cancel_date)
#             sheet.cell(row=row, column=7, value=invoice.cancel_time)
#             sheet.cell(row=row, column=8, value=invoice.cancel_period)
#             sheet.cell(row=row, column=9, value=invoice.cancel_reason)
#             sheet.cell(row=row, column=10, value=invoice.returntax_document_number)
#             sheet.cell(row=row, column=11, value=invoice.cancel_remark)
#             row += 1

#     # 匯出 Excel
#     output = BytesIO()
#     workbook.save(output)
#     output.seek(0)

#     response = HttpResponse(
#         output,
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="cancel.xlsx"'
#     return response





@csrf_exempt
def twb2bmainitem_update_cancel_status(request):
    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        invoice_id = data.get("invoice_id")
        cancel_reason = data.get("cancel_reason", "").strip()
        cancel_remark = data.get("cancel_remark", "").strip()

        if not invoice_id:
            return JsonResponse({"success": False, "message": "缺少發票 ID"}, status=400)
        if not cancel_reason:
            return JsonResponse({"success": False, "message": "請輸入作廢理由"}, status=400)

        try:
            invoice = TWB2BMainItem.objects.select_related('company').get(id=invoice_id)
        except TWB2BMainItem.DoesNotExist:
            return JsonResponse({"success": False, "message": "找不到該發票"}, status=404)

        if invoice.invoice_status == '未開立':
            return JsonResponse({"success": False, "message": "該發票為未開立狀態，無法作廢"}, status=400)
        if invoice.allowance_status == '已開立折讓單':
            return JsonResponse({"success": False, "message": "該發票已開立折讓單，無法作廢"}, status=400)
        if invoice.mof_response != 'S0001':
            return JsonResponse({"success": False, "message": "該發票稅局未認證，無法作廢"}, status=400)
        
        try:
            invoice_period = int(invoice.invoice_period)  # 確保是整數，如 11412
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "message": "發票期別資料異常"}, status=400)
        
        now = localtime()
        current_roc_year = now.year - 1911
        current_roc_month = now.month
        current_period = current_roc_year * 100 + current_roc_month  # e.g., 11502
        
        if invoice_period < current_period:
            if not invoice.returntax_document_number or invoice.returntax_document_number.strip() == '':
                return JsonResponse({
                    "success": False,
                    "message": "跨期作廢的發票需填寫專案作廢核准文號(returntax_document_number)。"
                }, status=400)


        invoice.cancel_date = now.date()
        invoice.cancel_time = now.time()
        invoice.cancel_reason = cancel_reason
        invoice.cancel_remark = cancel_remark
        invoice.invoice_status = "已作廢"
        invoice.save()


        
        # ✅ 檢查是否跨期作廢 → 要填 returntax_document_number


        # 產生 F0501 XML
        output_dir_F0501 = r"C:\Users\waylin\mydjango\e_invoice\F0501"
        xsd_path = r"C:\Users\waylin\mydjango\e_invoice\valid_xml\F0501.xsd"
        generate_F0501_xml_files(invoice, output_dir_F0501, xsd_path)

        # ✅ 發送作廢成功通知信
        to_email = "waylin@deloitte.com.tw"
        success_list = [f"{invoice.company.company_name} - {invoice.invoice_number}"]
        excluded_list = []
        send_invoice_canceled_email(to_email, success_count=1, excluded_count=0, success_list=success_list, excluded_list=excluded_list)

        invoice.save()
        
        return JsonResponse({"success": True, "message": "作廢成功並產出 F0501.xml"})

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "資料格式錯誤"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
    
    
    

# @csrf_exempt
# def twb2bmainitem_update_cancel_status(request):
#     if request.method != 'POST':
#         return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

#     try:
#         data = json.loads(request.body)
#         invoice_id = data.get("invoice_id")
#         void_reason = data.get("void_reason", "").strip()
#         void_remark = data.get("void_remark", "").strip()

#         if not invoice_id:
#             return JsonResponse({"success": False, "message": "缺少發票 ID"}, status=400)
#         if not void_reason:
#             return JsonResponse({"success": False, "message": "請輸入註銷理由"}, status=400)

#         try:
#             invoice = TWB2BMainItem.objects.select_related('company').get(id=invoice_id)
#         except TWB2BMainItem.DoesNotExist:
#             return JsonResponse({"success": False, "message": "找不到該發票"}, status=404)

#         if invoice.invoice_status == '未開立':
#             return JsonResponse({"success": False, "message": "該發票為未開立狀態，無法註銷"}, status=400)
#         if invoice.allowance_status == '已開立折讓單':
#             return JsonResponse({"success": False, "message": "該發票已開立折讓單，無法註銷"}, status=400)

#         now = localtime()
#         invoice.void_date = now.date()
#         invoice.void_time = now.time()
#         invoice.void_reason = void_reason
#         invoice.void_remark = void_remark
#         invoice.invoice_status = "已註銷"
#         invoice.save()

#         # 產生 F0701 XML
#         output_dir_F0701 = r"C:\Users\waylin\mydjango\e_invoice\F0701"
#         xsd_path = r"C:\Users\waylin\mydjango\e_invoice\valid_xml\F0701.xsd"
#         generate_F0701_xml_files(invoice, output_dir_F0701, xsd_path)

#         return JsonResponse({"success": True, "message": "註銷成功並產出 F0701.xml"})

#     except json.JSONDecodeError:
#         return JsonResponse({"success": False, "message": "資料格式錯誤"}, status=400)
#     except Exception as e:
#         return JsonResponse({"success": False, "message": str(e)}, status=500)

