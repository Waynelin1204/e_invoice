# ====== Python 標準函式庫 ======
import os
import json
import logging
import subprocess
from io import BytesIO
from datetime import datetime, timedelta
from collections import defaultdict
import xml.etree.ElementTree as ET
from django.utils.timezone import localtime
from decimal import Decimal, InvalidOperation
import random


# ====== 第三方套件 ======
import pandas as pd
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# ====== Django 基礎功能 ======
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.utils import timezone
from django.urls import reverse

# ====== Django DB 操作 ======
from django.db import connection
from django.db import transaction
from django.db.models import Q, Count, Sum

# ====== Django 使用者與驗證 ======
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test

# ====== 專案內部（Model 與 Form） ======
from e_invoices.models import (
    RegisterForm, LoginForm,
    Twa0101, Twa0101Item, Ocr, Ocritem, Company, UserProfile,
    NumberDistribution, TWB2BMainItem, TWB2BLineItem, TWAllowanceLineItem, TWAllowance,
    CompanyNotificationConfig, SystemConfig, TWB2BMainItemInS, TWB2BMainItemInE,
    TWB2BLineItemInS, TWB2BLineItemInE
)
from e_invoices.forms import NumberDistributionForm
from e_invoices.services import generate_invoice_B2C_a4, generate_invoice_B2B_a4,generate_invoice_B2B_format25_pdf
from e_invoices.services import (
    generate_F0501_xml_files,
    generate_F0401_xml_files, 
    generate_F0701_xml_files,
    generate_A0101_xml_files,
    generate_A0201_xml_files,
    generate_A0301_xml_files,
    generate_A0102_xml_files,
    generate_A0202_xml_files,
    generate_A0302_xml_files,
    )
from e_invoices.services import send_invoice_summary_email, generate_invoice_B2B_format25_pdf_stamp, send_invoice_canceled_email, send_number_low_storage_remind_email, send_insufficient_number_email, send_invoice_deleted_email
from decimal import Decimal, InvalidOperation

#aes_key = str(config.AES_KEY)
#aes_key = "73BE69AA9826613AFAEC11C215F08302"

#======================================================B2B存證模式進項發票明細=======================================================
@login_required 
def twb2blineitemins(request, id):
    document = get_object_or_404(TWB2BMainItemInS, id=id)
    items = document.items.all()  # 確保有正確查詢

    # 判斷是否有效
    print(f"原銷售金額: {document.sales_amount}")    
    print(f"原零稅銷售金額: {document.zerotax_sales_amount}")
    print(f"原免稅銷售金額: {document.freetax_sales_amount}")

    return render(request, 'document/twb2blineitemin_storage.html', {
        'document': document,
        'items': items,
    })

#======================================================B2B存證模式進項發票篩選=======================================================
def twb2bmainitemins_filter(request):
    # 預設顯示筆數
    display_limit = int(request.GET.get("display_limit", 20))  # 默認為20筆資料

    # 其他篩選條件
    invoice_status_filter = request.GET.get("invoice_status")
    # exchange_mode_filter = request.GET.get("exchange_mode")
    #void_status_filter = request.GET.get("void_status")
    tax_type_filter = request.GET.get("tax_type")
    company_id_filter = request.GET.get("company_id")  # 新增公司篩選條件


    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile

    # 取得該使用者可查看的公司名稱列表
    company_options = user_profile.viewable_companies.all()
    #viewable_company_codes = user_profile.viewable_companies.values_list('id', flat=True)
    viewable_company_codes = user_profile.viewable_companies.values_list('company_id', flat=True)



    # 計算兩個月前的日期
    two_months_ago = datetime.today() - timedelta(days=60)

    # 取得時間範圍的過濾條件（默認從兩個月前開始）
    start_date = request.GET.get("start_date", two_months_ago.strftime('%Y-%m-%d'))
    end_date = request.GET.get("end_date", datetime.today().strftime('%Y-%m-%d'))

    # 轉換成 datetime 格式
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    except ValueError:
        start_date = two_months_ago  # 如果格式錯誤，使用預設的兩個月前日期

    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        end_date = datetime.today()  # 如果格式錯誤，使用今天的日期

    if (end_date - start_date).days > 60:
        messages.error(request, "查詢區間不得超過 60 天")
        return redirect('twb2bmainitem')
    if start_date > end_date:
        messages.error(request, "開始日期不能大於結束日期")
        return redirect('twb2bmainitem')
    
    # 查詢條件構建
    filters = Q()
    if invoice_status_filter:
        filters &= Q(invoice_status=invoice_status_filter)
    
    #if void_status_filter:
    #    filters &= Q(void_status=void_status_filter)
    if tax_type_filter:
        filters &= Q(tax_type=tax_type_filter)
    if company_id_filter:
        filters &= Q(company_id=company_id_filter)

    # 限制在兩個月前到今天的時間範圍內
    filters &= Q(erp_date__range=[start_date, end_date])


    # 加入公司權限過濾條件
    filters &= Q(company__in=viewable_company_codes)

    # filters &= Q(exchange_mode=False)

    filters &= Q(buyer__exchange_mode=False)
       

    # 查詢所有符合條件的發票資料
    documents = TWB2BMainItemInS.objects.filter(filters).order_by('-erp_date')

    for document in documents:
        document.is_disabled = (
            document.invoice_status == '已作廢' or
            (
                document.tax_type == '2' and (
                    (document.zerotax_sales_amount or 0) <= 0 or
                    not document.customs_clearance_mark or
                    not document.bonded_area_confirm or
                    not document.zero_tax_rate_reason
                )
            )
        )


    # 分頁
    paginator = Paginator(documents, display_limit)  # 每頁顯示的資料筆數
    page_number = request.GET.get('page')  # 獲取當前頁碼
    page_obj = paginator.get_page(page_number)  # 根據頁碼獲取相應的頁面資料

    # 獲取篩選條件的選項
    invoice_status = TWB2BMainItemInS.objects.values_list('invoice_status', flat=True).distinct()
    #void_status = TWB2BMainItem.objects.values_list('void_status', flat=True).distinct()
    tax_type = TWB2BMainItemInS.objects.values_list('tax_type', flat=True).distinct()

    # 檢查公司ID篩選是否有效
    if company_id_filter and company_id_filter not in viewable_company_codes:
        messages.error(request, "您無權限查看該公司資料")
        return redirect('twb2bmainitem')
    

    return render(request, "twb2bmainitemin_storage_filter.html", {
        "company_id_filter": company_id_filter,
        "documents": page_obj,  # 傳遞分頁結果
        "company_options": company_options,
        "invoice_status": invoice_status,
        "tax_type_filter": tax_type_filter,
        "invoice_status_filter": invoice_status_filter,
        #"void_status": void_status,
        "tax_type": tax_type,
        "display_limit": display_limit,  # 傳遞選擇的筆數
        "start_date": start_date.strftime('%Y-%m-%d'),  # 顯示篩選的開始日期
        "end_date": end_date.strftime('%Y-%m-%d'),  # 顯示篩選的結束日期
    })

#======================================================B2B存證模式進項發票匯出=======================================================
@csrf_exempt
def twb2bmainitemins_export_invoices(request):
    if request.method != 'POST':
        return HttpResponse("Only POST allowed", status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    if not selected_ids:
        return HttpResponse("No invoice IDs provided", status=400)

    invoices = TWB2BMainItemInS.objects.filter(id__in=selected_ids).prefetch_related('items')
    if not invoices.exists():
        return HttpResponse("No invoices found", status=404)
    
    # 1️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'E0502.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 2️⃣ 寫入 Excel
    for invoice in invoices:
        for item in invoice.items.all():
            sheet.cell(row=row, column=1, value=invoice.company.company_id)
            sheet.cell(row=row, column=2, value=invoice.invoice_number)
            sheet.cell(row=row, column=3, value=invoice.invoice_date)
            sheet.cell(row=row, column=4, value=invoice.invoice_time)
            sheet.cell(row=row, column=5, value=invoice.invoice_type)
            sheet.cell(row=row, column=6, value=invoice.company.company_identifier)
            sheet.cell(row=row, column=7, value=invoice.buyer_identifier)
            sheet.cell(row=row, column=8, value=invoice.buyer_name)
            sheet.cell(row=row, column=9, value=invoice.buyer_bp_id)
            sheet.cell(row=row, column=10, value=invoice.buyer_remark)
            sheet.cell(row=row, column=11, value=invoice.main_remark)
            sheet.cell(row=row, column=12, value=invoice.customs_clearance_mark)
            sheet.cell(row=row, column=13, value=invoice.category)
            sheet.cell(row=row, column=14, value=invoice.relate_number)
            sheet.cell(row=row, column=15, value=invoice.bonded_area_confirm)
            sheet.cell(row=row, column=16, value=invoice.zero_tax_rate_reason)
            sheet.cell(row=row, column=17, value=invoice.reserved1)
            sheet.cell(row=row, column=18, value=invoice.reserved2)
            sheet.cell(row=row, column=19, value=invoice.sales_amount)
            sheet.cell(row=row, column=20, value=invoice.freetax_sales_amount)
            sheet.cell(row=row, column=21, value=invoice.zerotax_sales_amount)
            sheet.cell(row=row, column=22, value=invoice.tax_type)
            sheet.cell(row=row, column=23, value=invoice.tax_rate)
            sheet.cell(row=row, column=24, value=float(invoice.total_tax_amount))
            sheet.cell(row=row, column=25, value=float(invoice.total_amount))
            sheet.cell(row=row, column=26, value=invoice.original_currency_amount)
            sheet.cell(row=row, column=27, value=invoice.exchange_rate)
            sheet.cell(row=row, column=28, value=invoice.currency)
            sheet.cell(row=row, column=29, value=item.line_description)
            sheet.cell(row=row, column=30, value=item.line_quantity)
            sheet.cell(row=row, column=31, value=item.line_unit)
            sheet.cell(row=row, column=32, value=float(item.line_unit_price))
            sheet.cell(row=row, column=33, value=item.line_tax_type)
            sheet.cell(row=row, column=34, value=float(item.line_amount))
            sheet.cell(row=row, column=35, value=item.line_sequence_number)
            sheet.cell(row=row, column=36, value=item.line_remark)
            sheet.cell(row=row, column=37, value=item.line_relate_number)
            row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
    return response

#======================================================B2B交換模式進項發票開立確認(A0102)=======================================================
@csrf_exempt
def twb2bmainitemine_confirm(request):
    config = SystemConfig.objects.first()
    to_email = config.operator_output_email_address if config else None
    A0102_XSD_path = config.A0102_XSD_path

    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        invoice_id = data.get("invoice_id")
        # cancel_reason = data.get("cancel_reason", "").strip()
        # cancel_remark = data.get("cancel_remark", "").strip()

        if not invoice_id:
            return JsonResponse({"success": False, "message": "缺少發票 ID"}, status=400)  
        try:
            invoice = TWB2BMainItemInE.objects.select_related('company').get(id=invoice_id)
        except TWB2BMainItemInE.DoesNotExist:
            return JsonResponse({"success": False, "message": "找不到該發票"}, status=404)
        try:
            invoice_period = int(invoice.invoice_period)  # 確保是整數，如 11412
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "message": "發票期別資料異常"}, status=400)
            
        
        now = localtime()
        current_roc_year = now.year - 1911
        current_roc_month = now.month
        current_period = current_roc_year * 100 + current_roc_month  # e.g., 11502
        
        if invoice_period < current_period:
            if not invoice.returntax_document_number or invoice.returntax_document_number.strip() == '':
                return JsonResponse({
                    "success": False,
                    "message": "非本期發票"
                }, status=400)
            
        invoice.save()


        
        # ✅ 檢查是否跨期作廢 → 要填 returntax_document_number


        # 產生 A0102 XML
        configs = CompanyNotificationConfig.objects.filter(company__company_id=invoice.company.company_id).first()
        output_dir_A0102 = configs.output_dir_A0102
        generate_A0102_xml_files(invoice, output_dir_A0102, A0102_XSD_path)

        # ✅ 發送作廢成功通知信
        success_list = [f"{invoice.company.company_name} - {invoice.invoice_number}"]
        excluded_list = []
        #send_invoice_canceled_email(to_email, success_count=1, excluded_count=0, success_list=success_list, excluded_list=excluded_list)

        invoice.save()
        
        return JsonResponse({"success": True, "message": "已確認進項發票，產出 A0102.xml"})

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "資料格式錯誤"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

#======================================================B2B交換模式進項發票開立退回(A0301)=======================================================   
@csrf_exempt
def twb2bmainitemine_reject(request):
    config = SystemConfig.objects.first()
    to_email = config.operator_output_email_address if config else None
    A0301_XSD_path = config.A0301_XSD_path

    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        invoice_id = data.get("invoice_id")
        # cancel_reason = data.get("cancel_reason", "").strip()
        # cancel_remark = data.get("cancel_remark", "").strip()

        if not invoice_id:
            return JsonResponse({"success": False, "message": "缺少發票 ID"}, status=400)  
        try:
            invoice = TWB2BMainItemInE.objects.select_related('company').get(id=invoice_id)
        except TWB2BMainItemInE.DoesNotExist:
            return JsonResponse({"success": False, "message": "找不到該發票"}, status=404)
        try:
            invoice_period = int(invoice.invoice_period)  # 確保是整數，如 11412
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "message": "發票期別資料異常"}, status=400)
            
        
        now = localtime()
        current_roc_year = now.year - 1911
        current_roc_month = now.month
        current_period = current_roc_year * 100 + current_roc_month  # e.g., 11502
        
        if invoice_period < current_period:
            if not invoice.returntax_document_number or invoice.returntax_document_number.strip() == '':
                return JsonResponse({
                    "success": False,
                    "message": "非本期發票"
                }, status=400)
            
        invoice.save()


        
        # ✅ 檢查是否跨期作廢 → 要填 returntax_document_number


        # 產生 A0301 XML
        configs = CompanyNotificationConfig.objects.filter(company__company_id=invoice.company.company_id).first()
        output_dir_A0301 = configs.output_dir_A0301
        generate_A0301_xml_files(invoice, output_dir_A0301, A0301_XSD_path)

        # ✅ 發送作廢成功通知信
        success_list = [f"{invoice.company.company_name} - {invoice.invoice_number}"]
        excluded_list = []
        #send_invoice_canceled_email(to_email, success_count=1, excluded_count=0, success_list=success_list, excluded_list=excluded_list)

        invoice.save()
        
        return JsonResponse({"success": True, "message": "已退回進項發票，產出 A0301.xml"})

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "資料格式錯誤"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

#======================================================B2B交換模式進項發票作廢確認(A0202)=======================================================
@csrf_exempt
def twb2bmainitemine_cancel_confirm(request):
    config = SystemConfig.objects.first()
    to_email = config.operator_output_email_address if config else None
    A0202_XSD_path = config.A0202_XSD_path

    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
        invoice_id = data.get("invoice_id")
        # cancel_reason = data.get("cancel_reason", "").strip()
        # cancel_remark = data.get("cancel_remark", "").strip()

        if not invoice_id:
            return JsonResponse({"success": False, "message": "缺少發票 ID"}, status=400)  
        try:
            invoice = TWB2BMainItemInE.objects.select_related('company').get(id=invoice_id)
        except TWB2BMainItemInE.DoesNotExist:
            return JsonResponse({"success": False, "message": "找不到該發票"}, status=404)
        try:
            invoice_period = int(invoice.invoice_period)  # 確保是整數，如 11412
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "message": "發票期別資料異常"}, status=400)
            
        
        now = localtime()
        current_roc_year = now.year - 1911
        current_roc_month = now.month
        current_period = current_roc_year * 100 + current_roc_month  # e.g., 11502
        
        if invoice_period < current_period:
            if not invoice.returntax_document_number or invoice.returntax_document_number.strip() == '':
                return JsonResponse({
                    "success": False,
                    "message": "非本期發票"
                }, status=400)
            
        invoice.save()


        
        # ✅ 檢查是否跨期作廢 → 要填 returntax_document_number


        # 產生 A0301 XML
        configs = CompanyNotificationConfig.objects.filter(company__company_id=invoice.company.company_id).first()
        output_dir_A0202 = configs.output_dir_A0202
        generate_A0301_xml_files(invoice, output_dir_A0202, A0202_XSD_path)

        # ✅ 發送作廢成功通知信
        success_list = [f"{invoice.company.company_name} - {invoice.invoice_number}"]
        excluded_list = []
        #send_invoice_canceled_email(to_email, success_count=1, excluded_count=0, success_list=success_list, excluded_list=excluded_list)

        invoice.save()
        
        return JsonResponse({"success": True, "message": "已退回進項發票，產出 A0301.xml"})

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "資料格式錯誤"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
