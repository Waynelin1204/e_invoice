from django.shortcuts import render, redirect
import django.contrib.auth.views as auth_view
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.shortcuts import render
from e_invoices.models import RegisterForm
from e_invoices.models import LoginForm
from e_invoices.models import Invoice
from django.db.models import Count
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.http import JsonResponse
from e_invoices.models import Sapfagll03
from e_invoices.models import Myinvoiceportal 
from django.db import connection
from e_invoices.models import Twa0101
from e_invoices.models import Twa0101Item
from e_invoices.models import Ocritem
from e_invoices.models import Ocr
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_list_or_404
from django.shortcuts import render
from django.http import JsonResponse
import subprocess
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from e_invoices.models import Company
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from e_invoices.models import UserProfile, Company
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
import json
import os
from io import BytesIO
from django.conf import settings
from openpyxl import load_workbook
import logging
from datetime import datetime, timedelta
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib import messages
from e_invoices.models import NumberDistribution
from django.utils import timezone
from collections import defaultdict
from django.db import transaction
from django.forms import modelform_factory
from e_invoices.forms import NumberDistributionForm
from e_invoices.models import TWB2BMainItem
from e_invoices.models import TWB2BLineItem


logger = logging.getLogger(__name__)

@login_required		
def main(request):
    return render(request, 'main.html')

def upload_test(request):
    user_profile = request.user.profile
    
    # 取得該使用者可查看的公司名稱列表
    
    # 查詢符合條件的資料，並使用 prefetch_related 來查詢發票明細
    company_options = user_profile.viewable_companies.all()
    context = {
        "company_options": company_options,
    }

    return render(request, 'upload_test.html',context)

UPLOAD_DIR = "/home/pi/OCR/Samples"
@csrf_exempt
def upload_file(request):
    
    if request.method == "POST":
        if "file" not in request.FILES:
            return JsonResponse({"success":False, "error": "Didn't Receive"}, status=400)

        file = request.FIELS["file"]
        file_path = os.path.join(UPLOAD_DIR, file.name)

        try:
            with open(file_path, "wb") as f:
                 for chunk in file.chunks():
                     f.write(chunk)
		
            return JsonResponse({"success":True, "file_path": file_path})
        except Exception as e:
            return JsonResponse({"success":False, "error": str(e)}, status=500)
    return JsonResponse({"success":False, "error": "Invalid"}, status=400)



@csrf_exempt    
def run_script(request):
    """Execute the OCR script and return output as JSON."""
    if request.method == "POST":
        try:
            script_output = subprocess.check_output(["python3", "/home/pi/OCR/AWS_PARSE_multi.py"], text=True)
            return JsonResponse({"success": True, "output": script_output})
        except subprocess.CalledProcessError as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)



def front4(request):
	return render(request,'front4.html')

UPLOAD_DIR_TW = os.path.join(settings.BASE_DIR, "upload") 

@csrf_exempt
def upload_file_tw(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get("invoice_file")

        if not uploaded_file:
            return JsonResponse({"success": False, "error": "沒有收到檔案"}, status=400)

        os.makedirs(UPLOAD_DIR_TW, exist_ok=True)
        file_path = os.path.join(UPLOAD_DIR_TW, uploaded_file.name)

        try:
            with open(file_path, "wb") as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)

            return JsonResponse({"success": True, "file_path": file_path})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"success": False, "error": "無效的請求"}, status=400)

@csrf_exempt
def run_script_tw(request):
    """Execute the parse.py script which handles Excel to DB import."""
    parse_script_path = os.path.join(UPLOAD_DIR_TW, "import2sqlite.py")
    
    if request.method == "POST":
        try:
            # 執行 parse.py 腳本（Python 版本依實際情況調整：python 或 python3）
            script_output = subprocess.check_output(["python", parse_script_path], text=True)

            return JsonResponse({"success": True, "output": script_output})

        except subprocess.CalledProcessError as e:
            return JsonResponse({"success": False, "error": str(e.output)}, status=500)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)


def front4(request):
	return render(request,'front4.html')	
#def test(request):
#	return render(request,'test.html')

#@login_required	
#def twa0101(request):
#    # Fetch documents for display, with optional search filtering
#    documents = Twa0101.objects.all()
#    context = {
#	'documents': documents,
#    }
#    return render(request, 'test.html', context)


# 只有管理員可以訪問
def is_admin(user):
    return user.is_superuser or user.is_staff

@login_required
@user_passes_test(is_admin)
def manage_user_permissions(request):
    all_users = User.objects.all()
    companies = {company.id: f"{company.company_id} - {company.company_name}" for company in Company.objects.all()}  
    return render(request, "permissions.html", {"all_users": all_users, "companies": json.dumps(companies)})

# @login_required
# @user_passes_test(is_admin)
# def get_user_permissions(request, user_id):
    # user = get_object_or_404(User, id=user_id)

    # # **如果 `UserProfile` 不存在，則自動建立**
    # user_profile, _ = UserProfile.objects.get_or_create(user=user)

    # #viewable_companies = list(user_profile.viewable_companies.values_list("id", flat=True))
    # viewable_companies = list(Company.objects.values_list("id", flat=True))

    # return JsonResponse({"status": "exists", "viewable_companies": viewable_companies})

# @login_required
# @user_passes_test(is_admin)
# @csrf_exempt
# def update_permissions(request, user_id):
    # if request.method == "POST":
        # user = get_object_or_404(User, id=user_id)
        # user_profile, _ = UserProfile.objects.get_or_create(user=user)

        # selected_companies = request.POST.getlist("viewable_companies[]")  # 確保接收前端資料
        # selected_companies = list(map(int, selected_companies))  # 轉換為整數 ID

        # user_profile.viewable_companies.set(Company.objects.filter(id__in=selected_companies))
        # return JsonResponse({"status": "success", "message": "權限更新成功！"})

    # return JsonResponse({"status": "error", "message": "無效的請求"}, status=400)


@login_required
@user_passes_test(is_admin)
def get_user_permissions(request, user_id):
    try:
        user = get_object_or_404(User, id=user_id)

        # 確保 UserProfile 存在
        user_profile, _ = UserProfile.objects.get_or_create(user=user)

        # 取得該使用者已選取的公司
        viewable_companies = list(user_profile.viewable_companies.values_list("id", flat=True))

        return JsonResponse({"status": "exists", "viewable_companies": viewable_companies})

    except Exception as e:
        return JsonResponse({"status": "error", "message": f"發生錯誤: {str(e)}"})

# **更新使用者的可查看公司權限**
@csrf_exempt  # 這裡使用 @csrf_exempt 來免除 CSRF 檢查，視情況而定可選擇開啟 CSRF 檢查
@login_required
@user_passes_test(is_admin)
def update_permissions(request, user_id):
    if request.method == "POST":
        try:
            user = get_object_or_404(User, id=user_id)
            user_profile, _ = UserProfile.objects.get_or_create(user=user)

            # 解析前端傳過來的 JSON 資料
            data = json.loads(request.body)
            selected_companies = data.get("viewable_companies", [])

            # 確保傳入的 ID 是有效的
            valid_companies = Company.objects.filter(id__in=selected_companies)
            
            # 更新使用者的 viewable_companies 權限
            user_profile.viewable_companies.set(valid_companies)

            return JsonResponse({"status": "success", "message": "權限更新成功！"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": f"發生錯誤: {str(e)}"})

    return JsonResponse({"status": "error", "message": "無效的請求"}, status=400)


# 營業人管理
def company_detail(request):
    companies = Company.objects.all()
    return render(request, 'company_detail.html', {'companies': companies})    

# 營業人管理-檢視+修改
def company_detail_sub(request, company_id):
    print("Received company_id:", company_id)
    company = get_object_or_404(Company, company_id=company_id)
    head_offices = Company.objects.filter(company_type=0).exclude(id=company.id)

    if request.method == 'POST':
        # 抓取表單值
        company_register_name = request.POST.get('company_register_name', '').strip()
        company_identifier = request.POST.get('company_identifier', '').strip()
        company_name = request.POST.get('company_name', '').strip()
        company_address = request.POST.get('company_address', '').strip()
        company_type = request.POST.get('company_type')
        is_foreign_ecomm = request.POST.get('is_foreign_ecomm')
        tax_identifer = request.POST.get('tax_identifer', '').strip()
        email = request.POST.get('email', '').strip()
        reporting_period = request.POST.get('reporting_period')
        head_id = request.POST.get('head_company_identifer')

        errors = {}
        # 後端驗證
        if not re.fullmatch(r'^\d{8}$', company_identifier):
            errors['company_identifier'] = '請輸入8碼數字'
        
        if len(company_register_name) > 100:
            errors['company_register_name'] = '請輸入100碼以內字元'

        if len(company_name) > 100:
            errors['company_name'] = '請輸入100碼以內字元'

        if len(company_address) > 255:
            errors['company_address'] = '請輸入255碼以內字元'

        if not re.fullmatch(r'^\d{9}$', tax_identifer):
            errors['tax_identifer'] = '請輸入9碼數字'

        if email and not re.fullmatch(r'^.+@.+$', email):
            errors['email'] = '請輸入100碼以內含有「@」符號的電子郵件地址'

        if errors:
            head_offices = Company.objects.filter(company_type=0)
            form_data = request.POST.dict()
            return render(request, 'company_detail_sub.html', {
                'head_offices': head_offices,
                'errors': errors,
                'form_data': form_data
            })
        
        # 通過驗證才儲存
        company.company_register_name = company_register_name
        company.company_identifier = company_identifier
        company.company_name = company_name
        company.company_address = company_address
        company.company_type = int(company_type)
        company.is_foreign_ecomm = int(is_foreign_ecomm)
        company.tax_identifer = tax_identifer
        company.email = email
        company.reporting_period = reporting_period

        if head_id:
            company.head_company_identifer = Company.objects.get(id=head_id)
        else:
            company.head_company_identifer = None

        company.save()

        messages.success(request, '營業人資料儲存成功！')
        return redirect('company_detail_sub', company_id=company_id)

    return render(request, 'company_detail_sub.html', {
        'company': company,
        'head_offices': head_offices,
    })

# 營業人管理 - 新增
def company_add(request):
    if request.method == 'POST':
        company_id = request.POST.get('company_id', '').strip()
        company_identifier = request.POST.get('company_identifier', '').strip()
        company_register_name = request.POST.get('company_register_name', '').strip()
        company_name = request.POST.get('company_name', '').strip()
        company_address = request.POST.get('company_address', '').strip()
        head_company_identifer_id = request.POST.get('head_company_identifer') or None
        company_type = request.POST.get('company_type')
        is_foreign_ecomm = request.POST.get('is_foreign_ecomm')
        tax_identifer = request.POST.get('tax_identifer', '').strip()
        email = request.POST.get('email', '').strip()
        reporting_period = request.POST.get('reporting_period')

        errors = {}
        # 後端驗證
        if not re.fullmatch(r'^[a-zA-Z0-9]{1,10}$', company_id):
            errors['company_id'] = '請輸入10碼以內字元，僅限英文大小寫或數字'
        elif Company.objects.filter(company_id=company_id).exists():
            errors['company_id'] = '營業人代碼已存在，請重新輸入'

        if not re.fullmatch(r'^\d{8}$', company_identifier):
            errors['company_identifier'] = '請輸入8碼數字'
        
        if len(company_register_name) > 100:
            errors['company_register_name'] = '請輸入100碼以內字元'

        if len(company_name) > 100:
            errors['company_name'] = '請輸入100碼以內字元'

        if len(company_address) > 255:
            errors['company_address'] = '請輸入255碼以內字元'

        if not re.fullmatch(r'^\d{9}$', tax_identifer):
            errors['tax_identifer'] = '請輸入9碼數字'

        if email and not re.fullmatch(r'^.+@.+$', email):
            errors['email'] = '請輸入100碼以內含有「@」符號的電子郵件地址'

        if errors:
            head_offices = Company.objects.filter(company_type=0)
            form_data = request.POST.dict()
            return render(request, 'company_add.html', {
                'head_offices': head_offices,
                'errors': errors,
                'form_data': form_data
            })

        # 若驗證通過，建立資料
        Company.objects.create(
            company_id=company_id,
            company_identifier=company_identifier,
            company_register_name=company_register_name,
            company_name=company_name,
            company_address=company_address,
            head_company_identifer_id=head_company_identifer_id,
            company_type=company_type,
            is_foreign_ecomm=is_foreign_ecomm,
            tax_identifer=tax_identifer,
            email=email,
            reporting_period=reporting_period,
        )
        messages.success(request, '營業人資料儲存成功！')
        return redirect('company_detail')

    head_offices = Company.objects.filter(company_type=0)
    return render(request, 'company_add.html', {
        'head_offices': head_offices,
        'form_data': {},
        'errors': {}
    })
    
def register(request):

    form = RegisterForm()

    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')

    context = {
        'form':form
    }
    return render(request, 'accounts/register.html', context)

def sign_in(request):

    form=LoginForm()
    #form = AuthenticationForm()

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
             login(request, user)
             return redirect('main')  # ✅ Corrected

    context = {
        'form': form
    }
    return render(request, 'login.html', context)

def logout(request):
	logout(request)
	return redirect('/login.html')


#======================================================B2B發票列表=======================================================    

@login_required
def twb2bmainitem(request):
    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile
    
    # 取得該使用者可查看的公司名稱列表
    viewable_company_codes = user_profile.viewable_companies.values_list('id', flat=True)
    
    # 計算從今天開始往前推的60天的日期
    sixty_days_ago = datetime.today() - timedelta(days=60)
    
    # 篩選條件：只顯示最近60天的發票
    filter_conditions = Q(company__in=viewable_company_codes) & Q(erp_date__gte=sixty_days_ago)

    # 查詢符合條件的資料，並使用 prefetch_related 來查詢發票明細
    documents = TWB2BMainItem.objects.filter(filter_conditions).prefetch_related('items').order_by('-erp_date')
    company_options = user_profile.viewable_companies.all()


    # 分頁：每頁顯示25筆資料
    paginator = Paginator(documents, 25)  # 每頁顯示25筆資料
    page_number = request.GET.get('page')  # 取得當前頁數
    page_obj = paginator.get_page(page_number)  # 根據頁數取得對應的資料

    # 傳遞資料給模板
    context = {
        'documents': page_obj,  # 傳遞分頁後的資料
        "company_options": company_options,
    }
    print("🔍 可查看的公司 company_id：", list(viewable_company_codes))
    print("✅ 撈到的發票數：", TWB2BMainItem.objects.filter(filter_conditions).count())
    print("🟡 viewable_company_ids:", list(viewable_company_codes))
    print("🟡 篩選時間從:", sixty_days_ago)
    print("🟡 所有發票的公司代碼：", TWB2BMainItem.objects.values_list("company_code", flat=True).distinct())
    print("🟡 最近60天的發票：", TWB2BMainItem.objects.filter(erp_date__gte=sixty_days_ago).values_list("company_code", flat=True))
    print("🟡 完整符合條件的發票數：", TWB2BMainItem.objects.filter(filter_conditions).count())

    return render(request, 'twb2bmainitem.html', context)

#======================================================B2B發票明細=======================================================
    
def twb2blineitem(request, id):
    document = get_object_or_404(TWB2BMainItem, id=id)
    items = document.items.all()  # 確保有正確查詢
    return render(request, 'document/twb2blineitem.html', {'document': document, 'items': items})

#======================================================B2B發票篩選=======================================================

def twb2bmainitem_filter(request):
    # 預設顯示筆數
    display_limit = int(request.GET.get("display_limit", 20))  # 默認為20筆資料

    # 其他篩選條件
    invoice_status_filter = request.GET.get("invoice_status")
    void_status_filter = request.GET.get("void_status")
    tax_type_filter = request.GET.get("tax_type")
    company_id_filter = request.GET.get("company_id")  # 新增公司篩選條件


    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile

    # 取得該使用者可查看的公司名稱列表
    company_options = user_profile.viewable_companies.all()
    viewable_company_codes = user_profile.viewable_companies.values_list('id', flat=True)


    # 計算兩個月前的日期
    two_months_ago = datetime.today() - timedelta(days=60)

    # 取得時間範圍的過濾條件（默認從兩個月前開始）
    start_date = request.GET.get("start_date", two_months_ago.strftime('%Y-%m-%d'))
    end_date = request.GET.get("end_date", datetime.today().strftime('%Y-%m-%d'))

    # 轉換成 datetime 格式
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    except ValueError:
        start_date = two_months_ago  # 如果格式錯誤，使用預設的兩個月前日期

    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        end_date = datetime.today()  # 如果格式錯誤，使用今天的日期

    if (end_date - start_date).days > 60:
        messages.error(request, "查詢區間不得超過 60 天")
        return redirect('twb2bmainitem')
    if start_date > end_date:
        messages.error(request, "開始日期不能大於結束日期")
        return redirect('twb2bmainitem')
    
    # 查詢條件構建
    filters = Q()
    if invoice_status_filter:
        filters &= Q(invoice_status=invoice_status_filter)
    if void_status_filter:
        filters &= Q(void_status=void_status_filter)
    if tax_type_filter:
        filters &= Q(tax_type=tax_type_filter)
    if company_id_filter:
        filters &= Q(company__id=company_id_filter)

    # 限制在兩個月前到今天的時間範圍內
    filters &= Q(erp_date__range=[start_date, end_date])


    # 加入公司權限過濾條件
    filters &= Q(company__in=viewable_company_codes)

    # 查詢所有符合條件的發票資料
    invoices_list = TWB2BMainItem.objects.filter(filters).order_by('-erp_date')

    # 分頁
    paginator = Paginator(invoices_list, display_limit)  # 每頁顯示的資料筆數
    page_number = request.GET.get('page')  # 獲取當前頁碼
    page_obj = paginator.get_page(page_number)  # 根據頁碼獲取相應的頁面資料

    # 獲取篩選條件的選項
    invoice_status = TWB2BMainItem.objects.values_list('invoice_status', flat=True).distinct()
    void_status = TWB2BMainItem.objects.values_list('void_status', flat=True).distinct()
    tax_type = TWB2BMainItem.objects.values_list('tax_type', flat=True).distinct()

    # 檢查公司ID篩選是否有效
    if company_id_filter and int(company_id_filter) not in viewable_company_codes:
        messages.error(request, "您無權限查看該公司資料")
        return redirect('twb2bmainitem')

    return render(request, "twb2bmainitem.html", {
        "company_id_filter": company_id_filter,
        "documents": page_obj,  # 傳遞分頁結果
        "company_options": company_options,
        
        
        "invoice_status": invoice_status,
        "void_status": void_status,
        "tax_type": tax_type,
        "display_limit": display_limit,  # 傳遞選擇的筆數
        "documents": page_obj,  # 傳遞分頁結果
        "start_date": start_date.strftime('%Y-%m-%d'),  # 顯示篩選的開始日期
        "end_date": end_date.strftime('%Y-%m-%d'),  # 顯示篩選的結束日期
    })

#======================================================B2B發票匯出=======================================================

@csrf_exempt
def twb2bmainitem_export_invoices(request):
    if request.method != 'POST':
        return HttpResponse("Only POST allowed", status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    if not selected_ids:
        return HttpResponse("No invoice IDs provided", status=400)

    invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items')
    if not invoices.exists():
        return HttpResponse("No invoices found", status=404)

    # 1️⃣ 統計各公司所需發票數
    invoice_count_by_company_code = defaultdict(int)
    for invoice in invoices:
        #invoice_count_by_company_code[invoice.company_id] += 1  # invoice.company_id 是字串
        invoice_count_by_company_code[invoice.company.company_id] += 1
    # 2️⃣ 建立公司代碼對應的 Company 資料（查主鍵）
    company_map = {
        company.company_id: company for company in Company.objects.filter(company_id__in=invoice_count_by_company_code.keys())
    }

    # 3️⃣ 驗證每間公司是否有足夠的號碼可以使用
    insufficient_companies = []

    for company_code, count_needed in invoice_count_by_company_code.items():
        company_obj = company_map.get(company_code)
        if not company_obj:
            insufficient_companies.append(f"公司代碼 {company_code} 找不到對應公司資料")
            continue

        distributions = NumberDistribution.objects.filter(
            company=company_obj,
            status='available'
        )

        total_available = sum(
            int(d.end_number) - int(d.current_number or d.start_number) + 1
            for d in distributions
        )

        if total_available < count_needed:
            insufficient_companies.append(f"公司 {company_obj.company_name}（剩 {total_available} 張，需求 {count_needed} 張）")

    if insufficient_companies:
        return HttpResponse("號碼不足，請檢查以下公司：\n" + "\n".join(insufficient_companies), status=400)

    # 4️⃣ 準備號碼池（以 company.id 為 key）
    number_pool = defaultdict(list)
    for dist in NumberDistribution.objects.filter(status='available'):
        number_pool[dist.company.id].append(dist)

    # 5️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'A0401_Export.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 6️⃣ 開始配號與寫入 Excel
    with transaction.atomic():
        for invoice in invoices:
            #company_obj = company_map.get(invoice.company_id)
            company_obj = company_map.get(invoice.company.company_id)
            if not company_obj:
                #raise ValueError(f"找不到公司代碼為 {invoice.company_id} 的公司資料")
                raise ValueError(f"找不到公司代碼為 {invoice.company.company_id} 的公司資料")


            distributions = sorted(
                number_pool[company_obj.id],
                key=lambda d: int(d.current_number or d.start_number)
            )

            # 找一組號碼配給發票
            assigned = False
            for dist in distributions:
                current = int(dist.current_number or dist.start_number)
                if current <= int(dist.end_number):
                    invoice_number = f"{dist.initial_char}{str(current).zfill(len(dist.start_number))}"
                    invoice.invoice_number = invoice_number
                    invoice.invoice_status = '已開立'
                    invoice.invoice_date_time = timezone.now()
                    invoice.save()

                    dist.current_number = str(current + 1).zfill(len(dist.start_number))
                    dist.last_used_date = timezone.now().date()
                    dist.save()
                    assigned = True
                    break

            if not assigned:
                raise ValueError(f"公司 {company_obj.company_name} 號碼區間不足")

            # 寫入 Excel
            for item in invoice.items.all():
                sheet.cell(row=row, column=1, value=invoice.company.company_id)
                sheet.cell(row=row, column=2, value=invoice.invoice_number)
                sheet.cell(row=row, column=3, value=invoice.invoice_date)
                sheet.cell(row=row, column=4, value=invoice.invoice_time)
                #sheet.cell(row=row, column=5, value=invoice.company.company_name)
                sheet.cell(row=row, column=5, value=invoice.invoice_type)
                sheet.cell(row=row, column=6, value=invoice.company.company_identifier)
                sheet.cell(row=row, column=7, value=invoice.buyer_identifier)
                sheet.cell(row=row, column=8, value=invoice.buyer_name)
                sheet.cell(row=row, column=9, value=invoice.buyer_bp_id)
                sheet.cell(row=row, column=10, value=invoice.buyer_remark)
                sheet.cell(row=row, column=11, value=invoice.main_remark)
                sheet.cell(row=row, column=12, value=invoice.customs_clearance_mark)
                sheet.cell(row=row, column=13, value=invoice.category)
                sheet.cell(row=row, column=14, value=invoice.relate_number)
                sheet.cell(row=row, column=15, value=invoice.bonded_area_confirm)
                sheet.cell(row=row, column=16, value=invoice.zero_tax_rate_reason)
                sheet.cell(row=row, column=17, value=invoice.reserved1)
                sheet.cell(row=row, column=18, value=invoice.reserved2)
                sheet.cell(row=row, column=19, value=invoice.sales_amount)
                sheet.cell(row=row, column=20, value=invoice.freetax_sales_amount)
                sheet.cell(row=row, column=21, value=invoice.zerotax_sales_amount)
                sheet.cell(row=row, column=22, value=invoice.tax_type)
                sheet.cell(row=row, column=23, value=invoice.tax_rate)
                sheet.cell(row=row, column=24, value=float(invoice.total_tax_amount))
                sheet.cell(row=row, column=25, value=float(invoice.total_amount))
                sheet.cell(row=row, column=26, value=invoice.original_currency_amount)
                sheet.cell(row=row, column=27, value=invoice.exchange_rate)
                sheet.cell(row=row, column=28, value=invoice.currency)
                sheet.cell(row=row, column=29, value=item.line_description)
                sheet.cell(row=row, column=30, value=item.line_quantity)
                sheet.cell(row=row, column=31, value=item.line_unit)
                sheet.cell(row=row, column=32, value=float(item.line_unit_price))
                sheet.cell(row=row, column=33, value=item.line_tax_type)
                sheet.cell(row=row, column=34, value=float(item.line_amount))
                sheet.cell(row=row, column=35, value=item.line_sequence_number)
                sheet.cell(row=row, column=36, value=item.line_remark)
                sheet.cell(row=row, column=37, value=item.line_relate_number)
                row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
    return response


@csrf_exempt
def twb2bmainitem_delete_selected_invoices(request):
    if request.method == 'POST':
        selected_ids_raw = request.POST.get('selected_documents', '')
        selected_ids = selected_ids_raw.split(',') if selected_ids_raw else []

        if not selected_ids:
            messages.error(request, "刪除失敗：未選擇任何發票。")
            return redirect('test')

        deleted_count, _ = Twa0101.objects.filter(id__in=selected_ids).delete()

        if deleted_count > 0:
            messages.success(request, f"成功刪除 {deleted_count} 筆發票。")
        else:
            messages.warning(request, "未找到對應的發票資料，未進行刪除。")

        return redirect('twb2bmainitem')
    else:
        return redirect('twb2bmainitem')
        
@csrf_exempt
def twb2bmainitem_update_void_status(request):
    if request.method != 'POST':
        return HttpResponse("Only POST allowed", status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    if not selected_ids:
        return HttpResponse("No invoice IDs provided", status=400)

    invoices = TWB2BMainItem.objects.filter(id__in=selected_ids).prefetch_related('items')
    if not invoices.exists():
        return HttpResponse("No invoices found", status=404)

    # 5️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'A0401_Void.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 6️⃣ 開始配號與寫入 Excel
    with transaction.atomic():
        for invoice in invoices:

            # 寫入 Excel
            for item in invoice.items.all():
                #sheet.cell(row=row, column=1, value=invoice.company.company_name)
                sheet.cell(row=row, column=1, value=invoice.invoice_number)
                sheet.cell(row=row, column=2, value=invoice.invoice_date)
                sheet.cell(row=row, column=3, value=invoice.invoice_time)
                #sheet.cell(row=row, column=5, value=invoice.company.company_name)
                sheet.cell(row=row, column=4, value=invoice.invoice_type)
                sheet.cell(row=row, column=5, value=invoice.company.company_identifier)
                sheet.cell(row=row, column=6, value=invoice.seller_bp_id)
                sheet.cell(row=row, column=7, value=invoice.buyer_tax_id)
                sheet.cell(row=row, column=8, value=invoice.buyer_name)
                sheet.cell(row=row, column=9, value=invoice.buyer_bp_id)
                sheet.cell(row=row, column=10, value=invoice.buyer_remark)
                sheet.cell(row=row, column=11, value=invoice.main_remark)
                sheet.cell(row=row, column=12, value=invoice.customs_clearance_mark)
                sheet.cell(row=row, column=13, value=invoice.category)
                sheet.cell(row=row, column=14, value=invoice.relate_number)
                sheet.cell(row=row, column=15, value=invoice.bonded_area_confirm)
                sheet.cell(row=row, column=16, value=invoice.zero_tax_rate_reason)
                sheet.cell(row=row, column=17, value=invoice.reserved1)
                sheet.cell(row=row, column=18, value=invoice.reserved2)
                sheet.cell(row=row, column=19, value=invoice.sales_amount)
                sheet.cell(row=row, column=20, value=invoice.freetax_sales_amount)
                sheet.cell(row=row, column=21, value=invoice.zerotax_sales_amount)
                sheet.cell(row=row, column=22, value=invoice.tax_type)
                sheet.cell(row=row, column=23, value=invoice.tax_rate)
                sheet.cell(row=row, column=24, value=float(invoice.total_tax_amount))
                sheet.cell(row=row, column=25, value=float(invoice.total_amount))
                sheet.cell(row=row, column=26, value=invoice.original_currency_amount)
                sheet.cell(row=row, column=27, value=invoice.exchange_rate)
                sheet.cell(row=row, column=28, value=invoice.currency)
                sheet.cell(row=row, column=29, value=item.line_description)
                sheet.cell(row=row, column=30, value=item.line_quantity)
                sheet.cell(row=row, column=31, value=item.line_unit)
                sheet.cell(row=row, column=32, value=float(item.line_unit_price))
                sheet.cell(row=row, column=33, value=item.line_tax_type)
                sheet.cell(row=row, column=34, value=float(item.line_amount))
                sheet.cell(row=row, column=35, value=item.line_sequence_number)
                sheet.cell(row=row, column=36, value=item.line_remark)
                sheet.cell(row=row, column=37, value=item.line_relate_number)
                row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Void.xlsx"'
    return response

# def update_void_status(request):
#     if request.method == 'POST':
#         # 獲取所有選擇的發票 id
#         selected_invoices = request.POST.getlist('selected_documents')  # ['id1', 'id2', ...]

#         void_statuses = {}  # 儲存發票 id 和對應的作廢狀態

#         # 從表單中獲取每張發票的作廢狀態
#         for invoice_id in selected_invoices:
#             void_status = request.POST.get(f'void_status_{invoice_id}')
#             void_statuses[invoice_id] = void_status

#         # 根據 id 更新每張發票的作廢狀態
#         for invoice_id, void_status in void_statuses.items():
#             try:
#                 invoice = Invoice.objects.get(id=invoice_id)
#                 invoice.void_status = void_status
#                 invoice.save()
#             except Invoice.DoesNotExist:
#                 # 處理找不到發票的情況
#                 pass

#         # 完成後可以重定向回發票列表頁面
#         return redirect('test')  # 假設更新後重定向回發票列表頁
    
def number_distribution(request):
    # 取得所有發票號碼的狀態
    numbers = NumberDistribution.objects.all()
    return render(request, 'number_distribution.html', {'numbers':numbers})




    

# NumberDistributionForm = modelform_factory(
#     NumberDistribution,
#     fields='__all__',  # 你也可以列出特定欄位如：['company', 'start_number', 'end_number']
# )

def create_number_distribution(request):
    if request.method == 'POST':
        form = NumberDistributionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('create_number_distribution')  # 或導向你要的成功頁面
        else:
            print(form.errors)
    else:

        form = NumberDistributionForm()
    form.fields['company'].queryset = Company.objects.all()  # <--- 必加

    return render(request, 'create_number_distribution.html', {'form': form})


@csrf_exempt
def get_next_invoice_number(distribution):
    try:
        if distribution.current_number:
            next_number = int(distribution.current_number) + 1
        else:
            next_number = int(distribution.start_number)

        if next_number > int(distribution.end_number):
            raise ValueError(f"號碼區間已用完：{distribution.initial_char}{distribution.end_number}")

        return str(next_number).zfill(len(distribution.start_number))

    except ValueError as e:
        raise e
    except Exception:
        raise ValueError("發號時發生非預期錯誤")


#-----------------------------------------------------------------------------------------------
@login_required
def twa0101(request):
    # 取得登入使用者的 UserProfile
    user_profile = request.user.profile
    
    # 取得該使用者可查看的公司名稱列表
    viewable_company_names = user_profile.viewable_companies.values_list('company_name', flat=True)
    
    # 計算從今天開始往前推的60天的日期
    sixty_days_ago = datetime.today() - timedelta(days=60)
    
    # 篩選條件：只顯示最近60天的發票
    filter_conditions = Q(seller_name__in=viewable_company_names) & Q(invoice_date__gte=sixty_days_ago)

    # 查詢符合條件的資料，並使用 prefetch_related 來查詢發票明細
    documents = Twa0101.objects.filter(filter_conditions).prefetch_related('items').order_by('-invoice_date')

    # 分頁：每頁顯示25筆資料
    paginator = Paginator(documents, 25)  # 每頁顯示25筆資料
    page_number = request.GET.get('page')  # 取得當前頁數
    page_obj = paginator.get_page(page_number)  # 根據頁數取得對應的資料

    # 傳遞資料給模板
    context = {
        'documents': page_obj,  # 傳遞分頁後的資料
    }

    return render(request, 'test.html', context)
    
def twa0101_detail(request, id):
    document = get_object_or_404(Twa0101, id=id)
    items = document.items.all()  # 確保有正確查詢
    return render(request, 'document/twa0101_detail.html', {'document': document, 'items': items})

def invoice_filter(request):
    # 預設顯示筆數
    display_limit = int(request.GET.get("display_limit", 20))  # 默認為20筆資料

    # 其他篩選條件
    invoice_status_filter = request.GET.get("invoice_status")
    void_status_filter = request.GET.get("void_status")
    tax_type_filter = request.GET.get("tax_type")

    # 計算兩個月前的日期
    two_months_ago = datetime.today() - timedelta(days=60)

    # 取得時間範圍的過濾條件（默認從兩個月前開始）
    start_date = request.GET.get("start_date", two_months_ago.strftime('%Y-%m-%d'))
    end_date = request.GET.get("end_date", datetime.today().strftime('%Y-%m-%d'))

    # 轉換成 datetime 格式
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    except ValueError:
        start_date = two_months_ago  # 如果格式錯誤，使用預設的兩個月前日期

    try:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        end_date = datetime.today()  # 如果格式錯誤，使用今天的日期

    if (end_date - start_date).days > 60:
        messages.error(request, "查詢區間不得超過 60 天")
        return redirect('test')
    if start_date > end_date:
        messages.error(request, "開始日期不能大於結束日期")
        return redirect('test')
    
    # 查詢條件構建
    filters = Q()
    if invoice_status_filter:
        filters &= Q(invoice_status=invoice_status_filter)
    if void_status_filter:
        filters &= Q(void_status=void_status_filter)
    if tax_type_filter:
        filters &= Q(tax_type=tax_type_filter)

    # 限制在兩個月前到今天的時間範圍內
    filters &= Q(invoice_date__range=[start_date, end_date])

    # 查詢所有符合條件的發票資料
    invoices_list = Twa0101.objects.filter(filters).order_by('-invoice_date')

    # 分頁
    paginator = Paginator(invoices_list, display_limit)  # 每頁顯示的資料筆數
    page_number = request.GET.get('page')  # 獲取當前頁碼
    page_obj = paginator.get_page(page_number)  # 根據頁碼獲取相應的頁面資料

    # 獲取篩選條件的選項
    invoice_status = Twa0101.objects.values_list('invoice_status', flat=True).distinct()
    void_status = Twa0101.objects.values_list('void_status', flat=True).distinct()
    tax_type = Twa0101.objects.values_list('tax_type', flat=True).distinct()

    return render(request, "test.html", {
        "invoice_status": invoice_status,
        "void_status": void_status,
        "tax_type": tax_type,
        "display_limit": display_limit,  # 傳遞選擇的筆數
        "documents": page_obj,  # 傳遞分頁結果
        "start_date": start_date.strftime('%Y-%m-%d'),  # 顯示篩選的開始日期
        "end_date": end_date.strftime('%Y-%m-%d'),  # 顯示篩選的結束日期
    })

@login_required(login_url="Login")
def reconcil(request):
	display_limit = int(request.GET.get("display_limit",25))
	
	tax_code_filter = request.GET.get("tax_code")
	document_type_filter = request.GET.get("document_type")
	document_number_filter = request.GET.get("search","")


	query = """
		SELECT
			s.document_number,
			s.tax_code,
			s.document_type,
			s.posting_date,
			s.amount_in_doc_curr,
			s.document_currency,
			s.amount_in_lc,
			s.local_currency,
			s.including_tax_amount,
			i.internalId,
			i.totalPayableAmount,
			i.totalExcludingTax,
			i.totalNetAmount
		FROM e_invoices_sapfagll03 s
		LEFT JOIN e_invoices_myinvoiceportal i
		ON s.document_number = SUBSTR(i.internalId, 1, LENGTH(i.internalId) - 4)
	"""
	
	filters = []
	if tax_code_filter:
		filters.append(f"s.tax_code = '{tax_code_filter}'")
	if document_type_filter:
		filters.append(f"s.document_type = '{document_type_filter}'")
	if document_number_filter:
		document_number_filter = document_number_filter.strip()
		filters.append(f"CAST(s.document_number AS TEXT) LIKE '%{document_number_filter}%'")
		
	
	if filters:
		query += " WHERE " + " AND ".join(filters)
		
	query += " ORDER BY s.posting_date DESC"
	query += f" LIMIT {display_limit}"
	print(document_number_filter)
	print("check")


	with connection.cursor() as cursor:
		cursor.execute(query)
		merged_data = [
			{
				"document_number": row[0],
				"tax_code": row[1],
				"document_type":row[2],
				"posting_date": row[3],
				"amount_in_doc_curr":row[4],
				"document_currency":row[5],
				"amount_in_lc":row[6],
				"local_currency":row[7],
				"including_tax_amount":row[8],
				"internalId":row[9],
				"totalPayableAmount":row[10],
				"totalExcludingTax":row[11],
				"totalNetAmount":row[12]
			}
			for row in cursor.fetchall()
		]
	
		cursor.execute('SELECT DISTINCT "Tax Code" FROM e_invoices_sapfagll03')
		tax_codes = [row[0] for row in cursor.fetchall()]
		
		cursor.execute('SELECT DISTINCT "Document Type" FROM e_invoices_sapfagll03')
		document_types = [row[0] for row in cursor.fetchall()]

	
	return render(request, "reconcil.html", {
		"merged_data": merged_data,
		"tax_codes":tax_codes,
		"document_types":document_types,
		"display_limit":display_limit,
		"search_query":document_number_filter,
	})	
	# sap_data = Sapfagll03.objects.all()
	
	# invoice_data = {invoice.internal_id_trimmed: invoice for invoice in Myinvoiceportal.objects.all()}
	
	# merged_data = []
	
	# for sap in sap_data:
	
		# invoice = invoice_data.get(sap.document_number)
		# merged_data.append({
			# "document_number":sap.document_number,
			# "tax_code":sap.tax_code,
			# "document_type":sap.document_type,
			# "posting_date": sap.posting_date,
			# "amount_in_doc_curr":sap.amount_in_doc_curr,
			# "document_currency":sap.document_currency,
			# "amount_in_lc":sap.amount_in_lc,
			# "local_currency":sap.local_currency,
			# "including_tax_amount":sap.including_tax_amount,
			# "internalId":invoice.internal_id if invoice else "N/A",
			# "totalPayableAmount":invoice.totalPayableAmount if invoice else "N/A",
			# "totalExcludingTax":invoice.totalExcludingTax if invoice else "N/A",
			# "totalNetAmount":invoice.totalNetAmount if invoice else "N/A",
			
		# })
		
	# return render(request, "reconcil.html", {"merged_data":merged_data})

def invoice_list(request):
    invoices = Ocr.objects.all()  # Fetch all invoices
    
    search_query = request.GET.get('search', '')
    if search_query:
        invoices = invoices.filter(invoice_number__icontains=search_query)
    return render(request, 'invoices/invoice_list.html', {'invoices': invoices})



def invoice_detail(request, invoice_id):
    invoice = Ocr.objects.get(id=invoice_id)  # Fetch a specific invoice
    return render(request, 'invoices/invoice_detail.html', {'invoice': invoice})

# def invoice_filter(request):
#         display_limit = int(request.GET.get("display_limit", 25))
#         invoice_status_filter = request.GET.get("invoice_status")
#         void_status_filter = request.GET.get("void_status")
#         tax_type_filter = request.GET.get("tax_type")
# 	# 獲取當前用戶的可查看公司名稱列表
#         user_profile = request.user.profile
#         viewable_company_names = user_profile.viewable_companies.values_list('company_name', flat=True)
#         viewable_company_names_str = ', '.join([f"'{name}'" for name in viewable_company_names])
#         query = """
# 		SELECT
# 			s.invoice_number,
# 			s.buyer_name,
# 			s.seller_name,
# 			s.invoice_date,
# 			s.tax_type,
# 			s.total_amount,
# 			s.payment_status,
# 			s.invoice_status,
# 			s.void_status

# 		FROM e_invoices_twa0101 s
# 	"""
	
#         filters = []
#         if invoice_status_filter:
#                 filters.append(f"s.invoice_status = '{invoice_status_filter}'")
#         if void_status_filter:
#                 filters.append(f"s.void_status = '{void_status_filter}'")
#         if tax_type_filter:
#                 filters.append(f"s.tax_type = '{tax_type_filter}'")
# 	#if document_number_filter:
# 	#	document_number_filter = document_number_filter.strip()
# 	#	filters.append(f"CAST(s.document_number AS TEXT) LIKE '%{document_number_filter}%'")
		
#         if viewable_company_names_str:
#             filters.append(f"s.seller_name IN ({viewable_company_names_str})")
	    
#         if filters:
#                 query += " WHERE " + " AND ".join(filters)
		
#         query += " ORDER BY s.invoice_date DESC"
#         query += f" LIMIT {display_limit}"


#         with connection.cursor() as cursor:
#                 cursor.execute(query)
#                 merged_data = [
#                         {
# 				"invoice_number": row[0],
# 				"buyer_name": row[1],
# 				"seller_name":row[2],
# 				"invoice_date": row[3],
# 				"tax_type":row[4],
# 				"total_amount":row[5],
# 				"payment_status":row[6],
# 				"invoice_status":row[7],
# 				"void_status":row[8]
# 			}
# 			for row in cursor.fetchall()
# 		]
	
#                 cursor.execute('SELECT DISTINCT "invoice_status" FROM e_invoices_twa0101')
#                 invoice_status = [row[0] for row in cursor.fetchall()]
		
#                 cursor.execute('SELECT DISTINCT "void_status" FROM e_invoices_twa0101')
#                 void_status = [row[0] for row in cursor.fetchall()]
		
#                 cursor.execute('SELECT DISTINCT "tax_type" FROM e_invoices_twa0101')
#                 tax_type = [row[0] for row in cursor.fetchall()]		

	
#         return render(request, "test.html", {
# 		"invoice_status":invoice_status,
# 		"void_status":void_status,
# 		"tax_type":tax_type,
# 		"display_limit":display_limit,
# 		"documents":merged_data,
# 	})	


def generate_pdf(request, document_id):
    try:
        document = Invoice.objects.get(document_id=str(document_id))
    except Invoice.DoesNotExist:
        return HttpResponse("Document not found.", status=404)

    # Remove or comment out the early return below:
    # return HttpResponse(f"Document found: {document.document_id}, Status: {document.status}")

    # Use the XML file path from the instance (document) rather than the model class
    xml_file_path = document.xml_file_path  # Correct way to access the stored file path

    # Open the XML file from the stored path
    try:
        with open(xml_file_path, "r", encoding="utf-8") as file:
            xml_data = file.read()
    except FileNotFoundError:
        return HttpResponse("XML file not found.", status=404)

    # Parse XML data
    root = ET.fromstring(xml_data)
    extracted_text = f"Document ID: {document.document_id}\nStatus: {document.status}"

    response = HttpResponse(content_type="application/pdf")
    #response["Content-Disposition"] = "inline"  # Show in browser instead of downloading
    response["Content-Disposition"] = f'attachment; filename="invoice_{document_id}.pdf'
    p = canvas.Canvas(response, pagesize=letter)
    #p.drawString(200, 750, "eDocument Details")
    # Draw text at the specified positions
    p.setFont("Helvetica", 12)  # Set font size

    # Header Section
    p.drawString(400, 730, f"{document.buyer_name}")
    p.drawString(400, 750, f"Invoice No.: {document.document_id}")
    p.drawString(400, 730, f"Issue Date: {document.issue_date}")
    
    #p.drawString(400, 710, f"Customer No.: {document.buyer_name}")

    # Buyer & Delivery Details
    p.drawString(100, 670, f"Buyer Name: {document.buyer_name}")
    p.drawString(100, 650, f"Buyer Address: {document.buyer_addr1}")
    p.drawString(100, 630, f"Buyer TIN: {document.buyer_tin_id}")
    p.drawString(100, 610, f"Buyer BRN: {document.buyer_brn_id}")

    p.drawString(400, 670, f"Delivery Address: {document.delivery_addr1}")

    # Industry Classification & Tax Details
    #p.drawString(100, 570, f"Industry Classification: {document.industry_classification}")
    p.drawString(400, 550, f"Taxable Amount: {document.taxable_amount}")
    p.drawString(400, 530, f"Tax Amount: {document.tax_amount}")
    #p.drawString(400, 510, f"Total Excl. Tax: {document.total_excl_tax}")
    #p.drawString(400, 490, f"Total Incl. Tax: {document.total_incl_tax}")
    p.drawString(400, 470, f"Prepaid Amount: {document.prepaid_amount}")
    p.drawString(400, 450, f"Amount to Pay: {document.payable_amount}")

    # Supplier Details
    p.drawString(100, 430, f"Supplier Name: {document.supplier_name}")
    p.drawString(100, 410, f"Supplier Address: {document.supplier_addr1}")
    p.drawString(100, 390, f"Supplier TIN ID: {document.supplier_tin_id}")
    p.drawString(100, 370, f"Supplier BRN ID: {document.supplier_brn_id}")

    # Tax Summary
    #p.drawString(400, 370, f"Tax Code: {document.tax_code}")
    p.drawString(400, 350, f"Taxable Amount: {document.taxable_amount}")
    p.drawString(400, 330, f"Tax Rate: {document.tax_rate}")
    p.drawString(400, 310, f"Tax Amount: {document.tax_amount}")

    # Item Details (Product Information Table)
    p.drawString(100, 280, "Article No.")
    p.drawString(200, 280, "Description")
    p.drawString(400, 280, "Quantity")
    p.drawString(450, 280, "Unit")
    p.drawString(500, 280, "Unit Price")
    p.drawString(550, 280, "Tax Rate")
    p.drawString(600, 280, "Net Amount")

    p.drawString(100, 260, f"{document.item_id}")
    p.drawString(200, 260, f"{document.item_name}")
    p.drawString(400, 260, f"{document.item_quantity}")
    #p.drawString(450, 260, f"{document.item_unit}")
    p.drawString(500, 260, f"{document.item_unit_price}")
    p.drawString(550, 260, f"{document.tax_rate}")
    p.drawString(600, 260, f"{document.line_ext_amount}")

    # Total Calculation
    #p.drawString(500, 220, f"Subtotal: {document.subtotal}")
    p.drawString(500, 200, f"Taxable Amount: {document.taxable_amount}")
    p.drawString(500, 180, f"Tax Amount: {document.tax_amount}")
    #p.drawString(500, 160, f"Total Excl. Tax: {document.total_excl_tax}")
    #p.drawString(500, 140, f"Total Incl. Tax: {document.total_incl_tax}")
    p.drawString(500, 120, f"Prepaid Amount: {document.prepaid_amount}")
    p.drawString(500, 100, f"Amount to Pay: {document.payable_amount}")

    # Additional Information
    p.drawString(100, 80, f"XML File Path: {xml_file_path}")

    p.save()

    return response
# @csrf_exempt 
# def export_invoices(request):
#     if request.method != 'POST':
#         return HttpResponse("Only POST allowed", status=405)

#     raw_ids = request.POST.get("selected_documents", "")
#     selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
#     if not selected_ids:
#         return HttpResponse("No invoice IDs provided", status=400)
    
#     # #✅ 先更新發票狀態
#     # updated_count = Twa0101.objects.filter(id__in=selected_ids).update(invoice_status='已開立')
#     # print(f"更新了 {updated_count} 筆發票")
        

#     # ✅ 查詢發票與明細資料
#     invoices = Twa0101.objects.filter(id__in=selected_ids).prefetch_related('items', 'company')
#     if not invoices.exists():
#         return HttpResponse("No invoices found", status=404)

#     # ✅ 驗證每間公司是否有足夠的號碼可以使用
#     invoice_count_by_company = defaultdict(int)
#     for invoice in invoices:
#         invoice_count_by_company[invoice.company_id] += 1

#     insufficient_companies = []

#     for company_id, count_needed in invoice_count_by_company.items():
#         distributions = NumberDistribution.objects.filter(
#             company_id=company_id,
#             status='available'
#         )

#         total_available = sum(
#             int(d.end_number) - int(d.current_number or d.start_number) + 1
#             for d in distributions
#         )

#         if total_available < count_needed:
#             company = Company.objects.get(id=company_id)
#             insufficient_companies.append(f"{company.name}（剩 {total_available} 張，需求 {count_needed} 張）")

#     if insufficient_companies:
#         return HttpResponse("號碼不足，請檢查以下公司：\n" + "\n".join(insufficient_companies), status=400)

#     # ✅ 準備分配號碼
#     number_pool = defaultdict(list)  # company_id -> [NumberDistribution objects]
#     for distribution in NumberDistribution.objects.filter(status='available'):
#         number_pool[distribution.company_id].append(distribution)

#     # ✅ 載入 Excel 樣板
#     template_path = os.path.join(settings.BASE_DIR, 'export', 'A0401_Export.xlsx')
#     workbook = load_workbook(template_path)
#     sheet = workbook.active

#     row = 2  # Excel 的起始列

#     with transaction.atomic():
#         for invoice in invoices:
#             company_id = invoice.company_id
#             distributions = sorted(
#                 number_pool[company_id],
#                 key=lambda d: int(d.current_number or d.start_number)
#             )

#             # 取得一組可用的號碼
#             assigned = False
#             for dist in distributions:
#                 current = int(dist.current_number or dist.start_number)
#                 if current <= int(dist.end_number):
#                     invoice_number = f"{dist.initial_char}{str(current).zfill(len(dist.start_number))}"
#                     invoice.invoice_number = invoice_number
#                     invoice.invoice_status = '已開立'
#                     invoice.invoice_date_time = timezone.now()
#                     invoice.save()
#                     dist.current_number = str(current + 1).zfill(len(dist.start_number))
#                     dist.last_used_date = timezone.now().date()
#                     dist.save()
#                     assigned = True
#                     break
#             if not assigned:
#                 raise ValueError(f"{invoice.company.name} 號碼區間不足")

#             invoice.save()

#             # ✅ 每張發票配一號，明細多筆只配一次
#             for item in invoice.items.all():
#                 sheet.cell(row=row, column=1, value=invoice.invoice_number)
#                 sheet.cell(row=row, column=2, value=f"{invoice.invoice_date} {invoice.invoice_time}")
#                 sheet.cell(row=row, column=3, value=invoice.corporate_id)
#                 sheet.cell(row=row, column=4, value=invoice.buyer_name)
#                 sheet.cell(row=row, column=6, value=invoice.main_remark)
#                 sheet.cell(row=row, column=7, value=invoice.customs_clearance_mark)
#                 sheet.cell(row=row, column=8, value=invoice.relate_number)
#                 sheet.cell(row=row, column=9, value=invoice.sales_amount)
#                 sheet.cell(row=row, column=10, value=invoice.tax_type)
#                 sheet.cell(row=row, column=11, value=invoice.tax_rate)
#                 sheet.cell(row=row, column=12, value=invoice.tax_amount)
#                 sheet.cell(row=row, column=13, value=invoice.total_amount)
#                 sheet.cell(row=row, column=14, value=item.product_name)
#                 sheet.cell(row=row, column=15, value=item.quantity)
#                 sheet.cell(row=row, column=16, value=item.unit)
#                 sheet.cell(row=row, column=17, value=float(item.unit_price))
#                 sheet.cell(row=row, column=18, value=float(item.amount))
#                 sheet.cell(row=row, column=19, value=item.remark)
#                 sheet.cell(row=row, column=20, value=invoice.relate_number)
#                 row += 1

#         # ✅ 更新發票狀態
#         Twa0101.objects.filter(id__in=selected_ids).update(invoice_status='已開立')

#     # ✅ 匯出 Excel
#     output = BytesIO()
#     workbook.save(output)
#     output.seek(0)

#     response = HttpResponse(
#         output,
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
#     return response
   
# def export_invoices(request):
#     if request.method == 'POST':

#         raw_ids = request.POST.get("selected_documents", "")
#         selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]

#         if not selected_ids:
#             return HttpResponse("No invoice IDs provided", status=400)
#         print(f"Selected invoice IDs: {selected_ids}")

#         # ✅ 先更新發票狀態
#         updated_count = Twa0101.objects.filter(id__in=selected_ids).update(invoice_status='已開立')
#         print(f"更新了 {updated_count} 筆發票")
        

#         # ✅ 查詢發票與明細資料
#         invoices = Twa0101.objects.filter(id__in=selected_ids).prefetch_related('items')
#         if not invoices.exists():
#             return HttpResponse("No invoices found", status=404)

#         # ✅ 載入 Excel 樣板
#         template_path = os.path.join(settings.BASE_DIR, 'export', 'A0401_Export.xlsx')
#         workbook = load_workbook(template_path)
#         sheet = workbook.active

#         row = 2  # 從第 2 列開始填資料
#         for invoice in invoices:
#             for item in invoice.items.all():
#                 sheet.cell(row=row, column=1, value=invoice.invoice_number)
#                 sheet.cell(row=row, column=2, value=f"{invoice.invoice_date} {invoice.invoice_time}")
#                 sheet.cell(row=row, column=3, value=invoice.corporate_id)
#                 sheet.cell(row=row, column=4, value=invoice.buyer_name)
#                 sheet.cell(row=row, column=6, value=invoice.main_remark)
#                 sheet.cell(row=row, column=7, value=invoice.customs_clearance_mark)
#                 sheet.cell(row=row, column=8, value=invoice.relate_number)
#                 sheet.cell(row=row, column=9, value=invoice.sales_amount)
#                 sheet.cell(row=row, column=10, value=invoice.tax_type)
#                 sheet.cell(row=row, column=11, value=invoice.tax_rate)
#                 sheet.cell(row=row, column=12, value=invoice.tax_amount)
#                 sheet.cell(row=row, column=13, value=invoice.total_amount)
#                 sheet.cell(row=row, column=14, value=item.product_name)
#                 sheet.cell(row=row, column=15, value=item.quantity)
#                 sheet.cell(row=row, column=16, value=item.unit)
#                 sheet.cell(row=row, column=17, value=float(item.unit_price))
#                 sheet.cell(row=row, column=18, value=float(item.amount))
#                 sheet.cell(row=row, column=19, value=item.remark)
#                 sheet.cell(row=row, column=20, value=invoice.relate_number)
#                 row += 1

#         # ✅ 匯出為 Excel 並回傳下載
#         output = BytesIO()
#         workbook.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
#         return response

@login_required(login_url="Login")
def document_list(request):
    # Count the total number of documents for each status
    status_counts = Invoice.objects.values('status').annotate(count=Count('status'))

    # Convert the status counts into a dictionary with proper keys
    status_dict = {status['status']: status['count'] for status in status_counts}

    # Ensure all statuses exist in the dictionary, default to 0 if missing
    all_statuses = ['created', 'error', 'in_progress', 'completed']
    status_dict = {status: status_dict.get(status, 0) for status in all_statuses}

    # Get the count of all documents
    all_documents_count = Invoice.objects.count()

    # Fetch documents for display, with optional search filtering
    documents = Invoice.objects.all()

    # Handling search query if provided
    search_query = request.GET.get('search', '')
    if search_query:
        documents = documents.filter(document_id__icontains=search_query)

    # Pass counts and documents to the template
    context = {
        'documents': documents,
        'status_dict': status_dict,
        'all_documents_count': all_documents_count,
    }

    return render(request, 'edocument.html', context)




# def update_invoice_status(request):
#     if request.method == 'POST':
#         selected_id = request.POST.get("selected_documents", "").split(",")

#         # 確保發票號碼有效
#         selected_id = [num.strip() for num in selected_id if num.strip()]
        
#         if selected_id:
#             updated_count = Twa0101.objects.filter(id__in=selected_id).update(invoice_status='已開立')
#             print(f"更新了 {updated_count} 筆發票")
#         else:
#             print("沒有選中的發票")

#     return redirect('test')  # 替換為您的發票列表頁面名稱

@csrf_exempt
def export_invoices(request):
    if request.method != 'POST':
        return HttpResponse("Only POST allowed", status=405)

    raw_ids = request.POST.get("selected_documents", "")
    selected_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    if not selected_ids:
        return HttpResponse("No invoice IDs provided", status=400)

    invoices = Twa0101.objects.filter(id__in=selected_ids).prefetch_related('items')
    if not invoices.exists():
        return HttpResponse("No invoices found", status=404)

    # 1️⃣ 統計各公司所需發票數
    invoice_count_by_company_code = defaultdict(int)
    for invoice in invoices:
        #invoice_count_by_company_code[invoice.company_id] += 1  # invoice.company_id 是字串
        invoice_count_by_company_code[invoice.company.company_id] += 1
    # 2️⃣ 建立公司代碼對應的 Company 資料（查主鍵）
    company_map = {
        company.company_id: company for company in Company.objects.filter(company_id__in=invoice_count_by_company_code.keys())
    }

    # 3️⃣ 驗證每間公司是否有足夠的號碼可以使用
    insufficient_companies = []

    for company_code, count_needed in invoice_count_by_company_code.items():
        company_obj = company_map.get(company_code)
        if not company_obj:
            insufficient_companies.append(f"公司代碼 {company_code} 找不到對應公司資料")
            continue

        distributions = NumberDistribution.objects.filter(
            company=company_obj,
            status='available'
        )

        total_available = sum(
            int(d.end_number) - int(d.current_number or d.start_number) + 1
            for d in distributions
        )

        if total_available < count_needed:
            insufficient_companies.append(f"公司 {company_obj.company_name}（剩 {total_available} 張，需求 {count_needed} 張）")

    if insufficient_companies:
        return HttpResponse("號碼不足，請檢查以下公司：\n" + "\n".join(insufficient_companies), status=400)

    # 4️⃣ 準備號碼池（以 company.id 為 key）
    number_pool = defaultdict(list)
    for dist in NumberDistribution.objects.filter(status='available'):
        number_pool[dist.company.id].append(dist)

    # 5️⃣ 載入 Excel 樣板
    template_path = os.path.join(settings.BASE_DIR, 'export', 'A0401_Export.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    row = 2  # Excel 開始列

    # 6️⃣ 開始配號與寫入 Excel
    with transaction.atomic():
        for invoice in invoices:
            #company_obj = company_map.get(invoice.company_id)
            company_obj = company_map.get(invoice.company.company_id)
            if not company_obj:
                #raise ValueError(f"找不到公司代碼為 {invoice.company_id} 的公司資料")
                raise ValueError(f"找不到公司代碼為 {invoice.company.company_id} 的公司資料")


            distributions = sorted(
                number_pool[company_obj.id],
                key=lambda d: int(d.current_number or d.start_number)
            )

            # 找一組號碼配給發票
            assigned = False
            for dist in distributions:
                current = int(dist.current_number or dist.start_number)
                if current <= int(dist.end_number):
                    invoice_number = f"{dist.initial_char}{str(current).zfill(len(dist.start_number))}"
                    invoice.invoice_number = invoice_number
                    invoice.invoice_status = '已開立'
                    invoice.invoice_date_time = timezone.now()
                    invoice.save()

                    dist.current_number = str(current + 1).zfill(len(dist.start_number))
                    dist.last_used_date = timezone.now().date()
                    dist.save()
                    assigned = True
                    break

            if not assigned:
                raise ValueError(f"公司 {company_obj.company_name} 號碼區間不足")

            # 寫入 Excel
            for item in invoice.items.all():
                sheet.cell(row=row, column=1, value=invoice.invoice_number)
                sheet.cell(row=row, column=2, value=f"{invoice.invoice_date} {invoice.invoice_time}")
                sheet.cell(row=row, column=3, value=invoice.corporate_id)
                sheet.cell(row=row, column=4, value=invoice.buyer_name)
                sheet.cell(row=row, column=6, value=invoice.main_remark)
                sheet.cell(row=row, column=7, value=invoice.customs_clearance_mark)
                sheet.cell(row=row, column=8, value=invoice.relate_number)
                sheet.cell(row=row, column=9, value=invoice.sales_amount)
                sheet.cell(row=row, column=10, value=invoice.tax_type)
                sheet.cell(row=row, column=11, value=invoice.tax_rate)
                sheet.cell(row=row, column=12, value=invoice.tax_amount)
                sheet.cell(row=row, column=13, value=invoice.total_amount)
                sheet.cell(row=row, column=14, value=item.product_name)
                sheet.cell(row=row, column=15, value=item.quantity)
                sheet.cell(row=row, column=16, value=item.unit)
                sheet.cell(row=row, column=17, value=float(item.unit_price))
                sheet.cell(row=row, column=18, value=float(item.amount))
                sheet.cell(row=row, column=19, value=item.remark)
                sheet.cell(row=row, column=20, value=invoice.relate_number)
                row += 1

    # 匯出 Excel
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
    return response

@csrf_exempt
def delete_selected_invoices(request):
    if request.method == 'POST':
        selected_ids_raw = request.POST.get('selected_documents', '')
        selected_ids = selected_ids_raw.split(',') if selected_ids_raw else []

        if not selected_ids:
            messages.error(request, "刪除失敗：未選擇任何發票。")
            return redirect('test')

        deleted_count, _ = Twa0101.objects.filter(id__in=selected_ids).delete()

        if deleted_count > 0:
            messages.success(request, f"成功刪除 {deleted_count} 筆發票。")
        else:
            messages.warning(request, "未找到對應的發票資料，未進行刪除。")

        return redirect('test')
    else:
        return redirect('test')